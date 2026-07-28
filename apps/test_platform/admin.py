import json
import copy
import mimetypes
from pathlib import Path
from urllib.parse import urlencode

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.core.exceptions import PermissionDenied, ValidationError
from django.http import FileResponse, Http404, HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html, format_html_join
from django.utils.safestring import mark_safe

from .models import (
    ApprovedKnowledgeEntry,
    ExecutionPlanArtifact,
    TestExecutionRun,
    TestIntentImport,
    TestCategory,
    TestPlanArtifact,
    TestResourceProfile,
    TestWorkflow,
)
from .generation_service import (
    queue_design_generation,
    queue_execution_plan_generation,
)
from .file_lock import nonblocking_file_lock


admin.site.site_header = "TestConductor"
admin.site.site_title = "TestConductor"
admin.site.index_title = "测试流程"
admin.site.has_permission = lambda request: True
for registered_model in tuple(admin.site._registry):
    if registered_model._meta.app_label == "auth":
        admin.site.unregister(registered_model)

_django_admin_get_urls = admin.site.get_urls


def _local_admin_urls():
    account_route_names = {"login", "logout", "password_change", "password_change_done"}
    return [
        route
        for route in _django_admin_get_urls()
        if getattr(route, "name", None) not in account_route_names
    ]


admin.site.get_urls = _local_admin_urls

TEST_CATEGORY_CHOICES = list(TestCategory.choices)
TEST_CATEGORY_LABELS = dict(TEST_CATEGORY_CHOICES)
TECHNIQUE_LABELS = {
    "positive": "正向 / 正常",
    "negative": "异常 / 非法",
    "boundary": "边界",
    "state_transition": "状态迁移",
    "recovery": "恢复",
    "permission": "权限",
    "idempotency": "幂等",
    "random": "随机组合",
}
EXECUTOR_CATEGORY = {
    "stagehand_agent": "ui",
    "http_api": "api",
    "database": "database",
    "performance": "performance",
    "tcp_port": "port",
}
UI_EXECUTOR_KINDS = frozenset({"stagehand_agent"})


def _approved_knowledge_catalog():
    """Return user-selectable approved knowledge without exposing storage details."""

    from .intent.contracts import ApprovedKnowledge

    documents = {}
    systems = {}
    titles = {}
    try:
        for entry in ApprovedKnowledgeEntry.objects.filter(
            status=ApprovedKnowledgeEntry.Status.APPROVED
        ):
            document = ApprovedKnowledge(
                scope_id=entry.scope_id,
                knowledge_id=entry.knowledge_id,
                version=entry.version,
                approval_id=entry.approval_id,
                approved_at=entry.approved_at.isoformat(),
                content=entry.content,
                content_hash=entry.content_hash,
            )
            documents[document.scope_id] = document
            systems[document.scope_id] = entry.system_id
            titles[document.scope_id] = entry.title
    except (TypeError, ValueError):
        return [], {}, {}, "当前有已发布知识无法通过安全校验，请联系平台管理员。"

    configured = str(
        getattr(settings, "TEST_PLATFORM_APPROVED_KNOWLEDGE_CATALOG", "") or ""
    ).strip()
    if not configured:
        return list(documents.values()), systems, titles, ""
    path = Path(configured)
    if not path.is_absolute():
        path = Path(settings.BASE_DIR) / path
    try:
        from .intent.knowledge import ApprovedKnowledgeSourceStore

        store = ApprovedKnowledgeSourceStore.from_json(path)
        imported = [
            store.approved_document(source.metadata.source_ref)
            for source in store.sources()
        ]
    except (OSError, TypeError, ValueError):
        return (
            list(documents.values()),
            systems,
            titles,
            "当前无法读取外部业务知识目录。",
        )
    for document in imported:
        documents.setdefault(document.scope_id, document)
        systems.setdefault(document.scope_id, store.system_id)
        titles.setdefault(document.scope_id, document.knowledge_id)
    return list(documents.values()), systems, titles, ""


def _knowledge_choice_label(document, title: str | None = None) -> str:
    return str(title or document.knowledge_id or "已审核业务知识").strip()


def _generation_progress_display(obj, *, compact: bool = False):
    progress = dict(getattr(obj, "generation_progress", None) or {})
    if not progress:
        return "-"
    try:
        percent = max(0, min(100, int(progress.get("percent") or 0)))
    except (TypeError, ValueError):
        percent = 0
    message = str(progress.get("message") or "正在处理").strip()
    compact_class = " tb-generation-progress--compact" if compact else ""
    return format_html(
        '<div class="tb-generation-progress{}" role="progressbar" '
        'aria-valuemin="0" aria-valuemax="100" aria-valuenow="{}">'
        '<div class="tb-generation-progress__track">'
        '<span style="width: {}%"></span></div>'
        '<div class="tb-generation-progress__text"><strong>{}%</strong><span>{}</span></div>'
        "</div>",
        compact_class,
        percent,
        percent,
        percent,
        message,
    )


def _example_download(filename: str, content_type: str) -> FileResponse:
    root = (Path(settings.BASE_DIR) / "examples" / "test_resources").resolve()
    path = (root / filename).resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise Http404("示例文件路径无效") from exc
    if not path.is_file():
        raise Http404("示例文件不存在")
    return FileResponse(
        path.open("rb"),
        as_attachment=True,
        filename=filename,
        content_type=content_type,
    )


class AuditPayloadAdminMixin:
    """Present machine handoff data without dumping raw JSON into review forms."""

    audit_payload_specs = ()

    def get_urls(self):
        opts = self.model._meta
        custom = [
            path(
                "<path:object_id>/audit/<str:field_name>.json",
                self.admin_site.admin_view(self.download_audit_payload),
                name=f"{opts.app_label}_{opts.model_name}_audit_payload",
            )
        ]
        return custom + super().get_urls()

    def download_audit_payload(self, request, object_id, field_name):
        specs = {field: label for field, label, _ in self.audit_payload_specs}
        if field_name not in specs:
            raise Http404("未知系统记录")
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("记录不存在")
        if not self.has_view_or_change_permission(request, obj):
            raise PermissionDenied
        payload = getattr(obj, field_name, None)
        response = HttpResponse(
            json.dumps(
                payload if payload is not None else {},
                ensure_ascii=False,
                indent=2,
                sort_keys=True,
                default=str,
            )
            + "\n",
            content_type="application/json; charset=utf-8",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{self.model._meta.model_name}-{obj.pk}-{field_name}.json"'
        )
        return response

    @admin.display(description="系统记录")
    def system_audit_summary(self, obj):
        if obj is None:
            return "保存后生成系统记录"
        opts = self.model._meta
        download_name = f"admin:{opts.app_label}_{opts.model_name}_audit_payload"
        rows = []
        for field_name, label, purpose in self.audit_payload_specs:
            payload = getattr(obj, field_name, None)
            if payload:
                encoded = json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    default=str,
                ).encode("utf-8")
                state = f"已保存 · {len(encoded) / 1024:.1f} KiB"
                action = format_html(
                    '<a class="tb-audit-download" href="{}">下载 JSON</a>',
                    reverse(download_name, args=[obj.pk, field_name]),
                )
            else:
                state = "尚未生成"
                action = "-"
            rows.append(
                format_html(
                    '<tr><th>{}</th><td data-label="系统用途">{}</td>'
                    '<td data-label="状态">{}</td><td data-label="原始数据">{}</td></tr>',
                    label,
                    purpose,
                    state,
                    action,
                )
            )
        identity = getattr(obj, "design_id", None) or getattr(obj, "plan_id", None) or "-"
        return format_html(
            '<div class="tb-audit-summary">'
            '<dl><div><dt>内容身份</dt><dd><code>{}</code> · v{}</dd></div>'
            '<div><dt>内容 Hash</dt><dd><code>{}</code></dd></div>'
            '<div><dt>当前状态</dt><dd>{}</dd></div></dl>'
            '<table><thead><tr><th>记录</th><th>系统用途</th><th>状态</th><th>原始数据</th></tr></thead>'
            '<tbody>{}</tbody></table></div>',
            identity,
            getattr(obj, "version", "-"),
            getattr(obj, "content_hash", "-") or "-",
            obj.get_status_display() if hasattr(obj, "get_status_display") else "-",
            format_html_join("", "{}", ((row,) for row in rows)),
        )


class SingleRecordActionAdmin:
    """Keep workflow commands on one record instead of the bulk-action menu."""

    actions = ("mark_selected", "unmark_selected", "delete_selected")
    change_form_template = "admin/test_platform/change_form.html"
    change_list_template = "admin/test_platform/change_list_base.html"
    page_title = None
    page_description = ""
    polling_statuses = ()

    def has_module_permission(self, request):
        return True

    def has_view_permission(self, request, obj=None):
        return True

    def has_add_permission(self, request):
        return True

    def has_change_permission(self, request, obj=None):
        return True

    def has_delete_permission(self, request, obj=None):
        return True

    def log_addition(self, request, obj, message):
        return None

    def log_change(self, request, obj, message):
        return None

    def log_deletions(self, request, queryset):
        return None

    def get_record_commands(self, request, obj):
        return []

    def _all_record_commands(self, request, obj):
        commands = list(self.get_record_commands(request, obj))
        if obj is not None and hasattr(obj, "is_marked"):
            commands.insert(
                0,
                {
                    "name": "_toggle_mark",
                    "label": "取消标记" if obj.is_marked else "标记为重要",
                    "handler": "run_toggle_mark",
                    "kind": "secondary",
                },
            )
        return commands

    @admin.display(boolean=True, description="标记", ordering="is_marked")
    def marked_status(self, obj):
        return obj.is_marked

    @admin.action(description="标记为重要")
    def mark_selected(self, request, queryset):
        count = queryset.update(is_marked=True)
        self.message_user(request, f"已标记 {count} 条记录。", messages.SUCCESS)

    @admin.action(description="取消标记")
    def unmark_selected(self, request, queryset):
        count = queryset.update(is_marked=False)
        self.message_user(request, f"已取消 {count} 条标记。", messages.SUCCESS)

    def run_toggle_mark(self, request, obj):
        obj.is_marked = not obj.is_marked
        obj.save(update_fields=("is_marked", "updated_at"))
        self.message_user(
            request,
            "已标记为重要。" if obj.is_marked else "已取消标记。",
            messages.SUCCESS,
        )
        return HttpResponseRedirect(request.path)

    def changelist_view(self, request, extra_context=None):
        context = {
            "title": self.page_title or self.model._meta.verbose_name_plural,
            "page_description": self.page_description,
            "generation_auto_refresh": bool(
                self.polling_statuses
                and self.get_queryset(request)
                .filter(status__in=self.polling_statuses)
                .exists()
            ),
        }
        if any(field.name == "is_marked" for field in self.model._meta.fields):
            context["marked_records_url"] = (
                reverse(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                )
                + "?marked=yes"
            )
            context["marked_records_count"] = self.model._default_manager.filter(
                is_marked=True
            ).count()
        context.update(extra_context or {})
        return super().changelist_view(request, context)

    def show_record_save(self, request, obj) -> bool:
        return True

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        obj = self.get_object(request, object_id) if object_id else None
        context = {
            "record_commands": self._all_record_commands(request, obj),
            "record_show_save": self.show_record_save(request, obj),
            "record_show_delete": bool(obj and self.has_delete_permission(request, obj)),
            "display_record_title": (
                getattr(obj, "title", "") or getattr(obj, "name", "")
            )
            if obj
            else "",
            "generation_auto_refresh": bool(
                obj
                and self.polling_statuses
                and obj.status in self.polling_statuses
            ),
        }
        context.update(extra_context or {})
        return super().changeform_view(request, object_id, form_url, context)

    def response_change(self, request, obj):
        for command in self._all_record_commands(request, obj):
            if command["name"] in request.POST:
                return getattr(self, command["handler"])(request, obj)
        return super().response_change(request, obj)


class MarkedRecordFilter(admin.SimpleListFilter):
    title = "标记状态"
    parameter_name = "marked"

    def lookups(self, request, model_admin):
        return (("yes", "已标记"),)

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.filter(is_marked=True)
        return queryset


class ExecutionPlanStateFilter(admin.SimpleListFilter):
    title = "运行状态"
    parameter_name = "execution_state"

    def lookups(self, request, model_admin):
        return (
            ("waiting", "等待运行"),
            ("generating", "生成中"),
            ("needs_regeneration", "需重新生成"),
        )

    def queryset(self, request, queryset):
        states = {
            "waiting": (
                ExecutionPlanArtifact.Status.REVIEW,
                ExecutionPlanArtifact.Status.APPROVED,
            ),
            "generating": (ExecutionPlanArtifact.Status.GENERATING,),
            "needs_regeneration": (
                ExecutionPlanArtifact.Status.BLOCKED,
                ExecutionPlanArtifact.Status.CHANGES,
                ExecutionPlanArtifact.Status.ERROR,
                ExecutionPlanArtifact.Status.SUPERSEDED,
            ),
        }
        selected = states.get(self.value())
        return queryset.filter(status__in=selected) if selected else queryset


class ApprovalStateTabsMixin:
    """Split approval queues from completed and unsuccessful records."""

    approval_state_param = "record_state"
    approval_state_default = "pending"
    approval_state_labels = {
        "pending": "待处理",
        "completed": "已完成",
        "failed": "未通过",
    }
    approval_state_statuses = {}

    @admin.display(description="名称", ordering="title")
    def approval_record_title(self, obj):
        return obj.title

    def _approval_state(self, request):
        if (
            request is not None
            and request.GET.get("marked") == "yes"
            and self.approval_state_param not in request.GET
        ):
            return None
        value = (
            request.GET.get(self.approval_state_param)
            if request is not None
            else None
        )
        return value if value in self.approval_state_statuses else self.approval_state_default

    def _is_change_view(self, request):
        return getattr(getattr(request, "resolver_match", None), "url_name", "") == (
            f"{self.model._meta.app_label}_{self.model._meta.model_name}_change"
        )

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if self._is_change_view(request):
            return queryset
        selected = self._approval_state(request)
        if selected is None:
            return queryset
        return queryset.filter(status__in=self.approval_state_statuses[selected])

    def changelist_view(self, request, extra_context=None):
        selected = self._approval_state(request)
        counts = {
            key: self.model._default_manager.filter(status__in=statuses).count()
            for key, statuses in self.approval_state_statuses.items()
        }
        tabs = []
        for key, label in self.approval_state_labels.items():
            query = request.GET.copy()
            query[self.approval_state_param] = key
            query.pop("p", None)
            tabs.append(
                {
                    "key": key,
                    "label": label,
                    "count": counts.get(key, 0),
                    "selected": key == selected,
                    "url": "?" + urlencode(query, doseq=True),
                }
            )
        context = {"approval_state_tabs": tabs}
        context.update(extra_context or {})
        return super().changelist_view(request, context)


class ApprovalStateParameterFilter(admin.SimpleListFilter):
    """Consume the tab query parameter without duplicating a sidebar filter."""

    title = "审批记录"
    parameter_name = "record_state"
    template = "admin/test_platform/empty_filter.html"

    def lookups(self, request, model_admin):
        return tuple(model_admin.approval_state_labels.items())

    def queryset(self, request, queryset):
        return queryset


class TestCategoryFilter(admin.SimpleListFilter):
    title = "测试分类"
    parameter_name = "test_category"

    def lookups(self, request, model_admin):
        return TEST_CATEGORY_CHOICES

    def queryset(self, request, queryset):
        selected = self.value()
        if not selected:
            return queryset
        field_name = (
            "allowed_channels"
            if issubclass(queryset.model, TestWorkflow)
            else "test_categories"
        )
        ids = [
            pk
            for pk, categories in queryset.values_list("pk", field_name)
            if selected in (categories or [])
        ]
        return queryset.filter(pk__in=ids)


class ExecutionPlanSourceFilter(admin.SimpleListFilter):
    title = "计划来源"
    parameter_name = "plan_source"

    def lookups(self, request, model_admin):
        return (
            (TestPlanArtifact.SourceKind.IMPORTED, "手动导入"),
            (TestPlanArtifact.SourceKind.GENERATED, "测试意图生成"),
        )

    def queryset(self, request, queryset):
        selected = self.value()
        if selected not in {
            TestPlanArtifact.SourceKind.IMPORTED,
            TestPlanArtifact.SourceKind.GENERATED,
        }:
            return queryset
        return queryset.filter(source_test_plan__source_kind=selected)


class ExecutionRunDateFilter(admin.SimpleListFilter):
    """Consume date-range parameters without rendering preset date links."""

    title = "运行日期范围"
    parameter_name = "date_from"
    template = "admin/test_platform/empty_filter.html"

    def __init__(self, request, params, model, model_admin):
        super().__init__(request, params, model, model_admin)
        if "date_to" in params:
            value = params.pop("date_to")
            self.used_parameters["date_to"] = value[-1]

    def lookups(self, request, model_admin):
        return (("_date_picker", ""),)

    def expected_parameters(self):
        return ("date_from", "date_to")

    def queryset(self, request, queryset):
        form = ExecutionRunDateFilterForm(
            {
                "date_from": self.used_parameters.get("date_from", ""),
                "date_to": self.used_parameters.get("date_to", ""),
            }
        )
        if not form.is_valid():
            return queryset
        date_from = form.cleaned_data["date_from"]
        date_to = form.cleaned_data["date_to"]
        if date_from:
            queryset = queryset.filter(started_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(started_at__date__lte=date_to)
        return queryset


class ExecutionRunStatusFilter(admin.SimpleListFilter):
    title = "执行状态"
    parameter_name = "result_state"

    status_groups = {
        "queued": (TestExecutionRun.Status.QUEUED,),
        "running": (TestExecutionRun.Status.RUNNING,),
        "passed": (TestExecutionRun.Status.PASSED,),
        "failed": (
            TestExecutionRun.Status.FAILED,
            TestExecutionRun.Status.ERROR,
            TestExecutionRun.Status.INCONCLUSIVE,
        ),
        "blocked": (TestExecutionRun.Status.BLOCKED,),
        "dry_run": (TestExecutionRun.Status.DRY_RUN,),
    }

    def lookups(self, request, model_admin):
        return (
            ("queued", "等待中"),
            ("running", "执行中"),
            ("passed", "成功"),
            ("failed", "失败"),
            ("blocked", "阻断"),
            ("dry_run", "预检"),
        )

    def queryset(self, request, queryset):
        statuses = self.status_groups.get(self.value())
        if statuses is None:
            return queryset
        return queryset.filter(status__in=statuses)


def _category_badges(categories):
    values = list(dict.fromkeys(str(item) for item in (categories or [])))
    if not values:
        return "-"
    badges = []
    if len(values) > 1:
        badges.append(
            format_html(
                '<span class="tb-category tb-category--compound">复合测试</span>'
            )
        )
    badges.extend(
        format_html(
            '<span class="tb-category tb-category--{}">{}</span>',
            value,
            TEST_CATEGORY_LABELS.get(value, value),
        )
        for value in values
    )
    return format_html_join(" ", "{}", ((badge,) for badge in badges))


def _design_categories(design):
    selections = design.get("selections") or {}
    configured = selections.get("allowed_channels") or []
    if configured:
        return list(dict.fromkeys(str(item) for item in configured))
    return list(
        dict.fromkeys(
            str(item.get("channel_hint"))
            for scenario in design.get("scenarios", [])
            for item in [
                *scenario.get("operations", []),
                *scenario.get("expected_results", []),
            ]
            if item.get("channel_hint")
        )
    )


def _plan_categories(plan):
    return list(
        dict.fromkeys(
            EXECUTOR_CATEGORY.get(
                str(stage.get("executor_kind") or ""),
                str(stage.get("executor_kind") or ""),
            )
            for flow in plan.get("flows", [])
            for stage in flow.get("stages", [])
            if stage.get("executor_kind") and _stage_has_operations(stage)
        )
    )


def _category_stage_counts(plan):
    counts = {}
    for flow in plan.get("flows", []):
        for stage in flow.get("stages", []):
            if not _stage_has_operations(stage):
                continue
            executor = str(stage.get("executor_kind") or "")
            category = EXECUTOR_CATEGORY.get(executor, executor)
            if category:
                counts[category] = counts.get(category, 0) + 1
    return counts


def _status_badge(value, label):
    return format_html(
        '<span class="tb-status tb-status--{}">{}</span>',
        str(value or "unknown"),
        label or value or "-",
    )

class TestResourceProfileForm(forms.ModelForm):
    """Select resource kinds first, then configure only their real inputs."""

    resource_types = forms.MultipleChoiceField(
        choices=TEST_CATEGORY_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=True,
        label="本配置包含的测试能力（可多选）",
        help_text="复合测试请一次勾选全部所需能力；它们共同属于同一个被测系统。",
    )
    api_base_url = forms.URLField(
        required=False,
        assume_scheme="https",
        label="被测 API 基础地址",
        help_text="例如 https://staging.example.test/api。",
    )

    class Meta:
        model = TestResourceProfile
        exclude = ("environment",)
        widgets = {
            "ui_agent_asset_text": forms.Textarea(attrs={"rows": 5}),
            "api_asset_text": forms.Textarea(attrs={"rows": 5}),
            "database_asset_text": forms.Textarea(attrs={"rows": 5}),
            "performance_asset_text": forms.Textarea(attrs={"rows": 5}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["system_id"].required = True
        friendly_fields = {
            "system_id": ("被测系统", "填写团队日常使用的系统名称或简称。"),
            "ui_agent_asset_file": (
                "网页 Agent 资料文件（可选）",
                "上传包含 URL、功能和最大步数的表格或常见文本文件。",
            ),
            "ui_agent_asset_text": (
                "网页 Agent 资料说明（可选）",
                "大致填写 URL、页面或功能名称、最大步数即可；与文件二选一。",
            ),
            "api_openapi_file": (
                "接口资料文件（可选）",
                "OpenAPI 可直接解析；Markdown、Word、JSON、YAML 等资料由模型整理。",
            ),
            "api_asset_text": (
                "接口资料说明（可选）",
                "大致写明 method、path、参数和用途即可；与文件二选一。",
            ),
            "database_query_file": (
                "数据库资料文件（可选）",
                "可上传 DDL、数据字典、表结构文档或现有策略；与文字说明二选一。",
            ),
            "database_asset_text": (
                "数据库资料说明（可选）",
                "大致写明数据库类型、可测试的表和字段即可，模型会整理结构。",
            ),
            "database_connection_ref": (
                "数据库连接引用",
                "选择运行环境中已登记的目标数据库名称；不要填写连接串、账号或密码。",
            ),
            "performance_profile_file": (
                "性能资料文件（可选）",
                "可上传现有配置、需求文档或历史方案；与文字说明二选一。",
            ),
            "performance_asset_text": (
                "性能资料说明（可选）",
                "大致写明目标、负载上限和关注指标即可，模型会整理配置。",
            ),
        }
        for field_name, (label, help_text) in friendly_fields.items():
            self.fields[field_name].label = label
            self.fields[field_name].help_text = help_text
        if self.instance.pk:
            self.initial["resource_types"] = sorted(
                self.instance.configured_channels()
            )

    def clean(self):
        cleaned = super().clean()
        selected = set(cleaned.get("resource_types") or [])
        if cleaned.get("ui_agent_asset_file") and str(
            cleaned.get("ui_agent_asset_text") or ""
        ).strip():
            self.add_error("ui_agent_asset_text", "网页 Agent 资料文件和文字说明选择一种即可")
        for file_name, text_name, label in (
            ("api_openapi_file", "api_asset_text", "接口资料"),
            ("database_query_file", "database_asset_text", "数据库资料"),
            ("performance_profile_file", "performance_asset_text", "性能资料"),
        ):
            if cleaned.get(file_name) and str(cleaned.get(text_name) or "").strip():
                self.add_error(text_name, f"{label}文件和文字说明选择一种即可")
        ui_complete = bool(
            cleaned.get("ui_agent_asset_file")
            or str(cleaned.get("ui_agent_asset_text") or "").strip()
        )
        complete = {
            "ui": ui_complete,
            "api": bool(
                (cleaned.get("api_openapi_file") or cleaned.get("api_asset_text"))
                and cleaned.get("api_base_url")
            ),
            "database": bool(
                (cleaned.get("database_query_file") or cleaned.get("database_asset_text"))
                and cleaned.get("database_connection_ref")
            ),
            "performance": bool(
                cleaned.get("performance_profile_file")
                or cleaned.get("performance_asset_text")
            ),
            "port": bool(
                cleaned.get("port_host") and cleaned.get("port_number") is not None
            ),
        }
        missing = [
            TEST_CATEGORY_LABELS[category]
            for category in selected
            if not complete[category]
        ]
        if missing:
            self.add_error(
                "resource_types",
                "以下资源尚未填写完整: " + "、".join(missing),
            )
        if not missing:
            from .planning.resources import validate_resource_source_inputs

            candidate = copy.copy(self.instance)
            for field_name in (
                "system_id",
                "environment",
                "ui_agent_asset_file",
                "ui_agent_asset_text",
                "api_openapi_file",
                "api_asset_text",
                "api_base_url",
                "database_query_file",
                "database_asset_text",
                "database_connection_ref",
                "performance_profile_file",
                "performance_asset_text",
            ):
                if field_name in cleaned:
                    setattr(candidate, field_name, cleaned.get(field_name))
            try:
                validate_resource_source_inputs(candidate)
            except (TypeError, ValueError) as exc:
                self.add_error("resource_types", str(exc))
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        selected = set(self.cleaned_data.get("resource_types") or [])
        category_fields = {
            "ui": (
                "ui_agent_asset_file",
                "ui_agent_asset_text",
            ),
            "api": ("api_openapi_file", "api_asset_text", "api_base_url"),
            "database": (
                "database_query_file",
                "database_asset_text",
                "database_connection_ref",
            ),
            "performance": ("performance_profile_file", "performance_asset_text"),
            "port": ("port_host", "port_number"),
        }
        for category, fields in category_fields.items():
            if category in selected:
                continue
            for field_name in fields:
                setattr(instance, field_name, None if field_name == "port_number" else "")
        if commit:
            instance.save()
        return instance


class ApprovedKnowledgeEntryForm(forms.ModelForm):
    """Import a document or paste one concise, reviewable knowledge item."""

    class Meta:
        model = ApprovedKnowledgeEntry
        fields = ("system_id", "title", "content", "source_file")
        widgets = {"content": forms.Textarea(attrs={"rows": 10})}

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Django removes readonly model fields from the generated form.  An
        # approved entry is entirely readonly, so keep the form usable when
        # the change page is rendered after publication.
        if "title" in self.fields:
            self.fields["title"].required = False
            self.fields["title"].help_text = "可不填；上传文件时将使用文件名。"
        if "content" in self.fields:
            self.fields["content"].required = False
            self.fields["content"].help_text = "可直接填写规则，也可以上传文件自动提取后再审核。"
        if "system_id" in self.fields:
            self.fields["system_id"].help_text = "必须与测试资源中的被测系统一致。"

    def clean(self):
        cleaned = super().clean()
        uploaded = cleaned.get("source_file")
        content = str(cleaned.get("content") or "").strip()
        if uploaded and not content:
            from .ingestion.adapters import parse_input_file
            from .ingestion.contracts import IngestionError, IngestionLimits, InputFile

            limits = IngestionLimits().validate()
            size = int(getattr(uploaded, "size", 0) or 0)
            if size > limits.max_file_bytes:
                self.add_error(
                    "source_file",
                    f"文件不能超过 {limits.max_file_bytes // (1024 * 1024)} MiB",
                )
                return cleaned
            data = uploaded.read(limits.max_file_bytes + 1)
            uploaded.seek(0)
            try:
                parsed = parse_input_file(
                    InputFile(
                        filename=Path(uploaded.name).name,
                        data=data,
                        content_type=getattr(uploaded, "content_type", None),
                    ),
                    limits,
                )
            except IngestionError as exc:
                self.add_error("source_file", str(exc))
                return cleaned
            content = "\n\n".join(
                str(value).strip()
                for _, value in parsed.requirements
                if str(value).strip()
            )
            cleaned["content"] = content
        if not content:
            self.add_error("content", "请填写知识内容，或上传一份可以提取文字的文件")
        if not str(cleaned.get("title") or "").strip() and uploaded:
            cleaned["title"] = Path(uploaded.name).stem[:200]
        if not str(cleaned.get("title") or "").strip():
            self.add_error("title", "请填写知识标题")
        return cleaned


class TestWorkflowForm(forms.ModelForm):
    """Human-friendly admin input; generated artifacts remain read-only."""

    knowledge_scope_ids = forms.MultipleChoiceField(
        choices=(),
        widget=forms.CheckboxSelectMultiple,
        required=False,
        label="可参考的业务知识（可选）",
        help_text="只选择与本次需求有关的已审核规则；未选择时仍可继续生成测试计划。",
    )

    allowed_channels = forms.MultipleChoiceField(
        choices=TEST_CATEGORY_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=True,
        label="测试分类（可多选）",
        help_text="测试计划会按所选分类生成场景，后续执行计划再分别转换为对应执行产物。",
    )

    class Meta:
        model = TestWorkflow
        fields = (
            "requirement_text",
            "requirement_file",
            "resource_profile",
            "allowed_channels",
            "knowledge_scope_ids",
        )
        widgets = {
            "requirement_text": forms.Textarea(attrs={"rows": 10}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        (
            documents,
            knowledge_systems,
            knowledge_titles,
            catalog_error,
        ) = _approved_knowledge_catalog()
        self.knowledge_scope_systems = knowledge_systems
        choices = [
            (
                document.scope_id,
                _knowledge_choice_label(
                    document,
                    knowledge_titles.get(document.scope_id),
                ),
            )
            for document in documents
        ]
        retained = list(getattr(self.instance, "knowledge_scope_ids", []) or [])
        if not choices and retained:
            choices = [(scope_id, "已选择的业务知识") for scope_id in retained]
            self.fields["knowledge_scope_ids"].disabled = True
        self.fields["knowledge_scope_ids"].choices = choices
        if catalog_error:
            self.fields["knowledge_scope_ids"].help_text = catalog_error
        elif not choices:
            self.fields["knowledge_scope_ids"].help_text = format_html(
                '当前没有已发布的业务知识。请先到<a href="{}">业务知识库</a>导入并发布；'
                "也可以暂不使用知识增强。",
                reverse("admin:test_platform_approvedknowledgeentry_changelist"),
            )
        if self.instance.pk:
            self.initial["allowed_channels"] = self.instance.allowed_channels or []
            self.initial["knowledge_scope_ids"] = retained
        if "resource_profile" in self.fields:
            self.fields["resource_profile"].queryset = TestResourceProfile.objects.filter(
                enabled=True
            )
            self.fields["resource_profile"].required = True

    def clean(self):
        cleaned = super().clean()
        from .ingestion import IngestionLimits

        allowed = set(cleaned.get("allowed_channels") or [])
        document_fields = ("requirement_file",)
        if not (
            cleaned.get("requirement_text")
            or any(cleaned.get(name) for name in document_fields)
        ):
            raise ValidationError("请填写需求原文或至少上传一份需求/测试说明文件")
        limits = IngestionLimits().validate()
        uploads = [cleaned.get(name) for name in document_fields if cleaned.get(name)]
        if len(uploads) > limits.max_files:
            raise ValidationError(f"最多上传 {limits.max_files} 份文件")
        total_bytes = 0
        for field_name in document_fields:
            uploaded = cleaned.get(field_name)
            if not uploaded:
                continue
            size = int(getattr(uploaded, "size", 0) or 0)
            total_bytes += size
            if size > limits.max_file_bytes:
                self.add_error(
                    field_name,
                    f"单个文件不能超过 {limits.max_file_bytes // (1024 * 1024)} MiB",
                )
        if total_bytes > limits.max_total_file_bytes:
            raise ValidationError(
                f"全部文件合计不能超过 {limits.max_total_file_bytes // (1024 * 1024)} MiB"
            )
        profile = cleaned.get("resource_profile")
        if profile is not None and not profile.enabled:
            raise ValidationError("所选测试资源配置已停用")
        if profile is not None:
            unsupported = sorted(
                channel
                for channel in allowed
                if channel not in profile.configured_channels()
            )
            if unsupported:
                raise ValidationError(
                    "测试资源配置不支持所选测试分类: "
                    + "、".join(TEST_CATEGORY_LABELS.get(item, item) for item in unsupported)
                )
        selected_knowledge = list(cleaned.get("knowledge_scope_ids") or [])
        mismatched = [
            scope_id
            for scope_id in selected_knowledge
            if profile is not None
            and self.knowledge_scope_systems.get(scope_id)
            and self.knowledge_scope_systems[scope_id] != profile.system_id
        ]
        if mismatched:
            self.add_error(
                "knowledge_scope_ids",
                "所选业务知识不属于当前被测系统，请更换测试资源或取消选择。",
            )
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        instance.coverage_by_category = {}
        instance.allowed_channels = list(self.cleaned_data.get("allowed_channels") or [])
        instance.knowledge_scope_ids = list(
            self.cleaned_data.get("knowledge_scope_ids") or []
        )
        profile = self.cleaned_data.get("resource_profile")
        if profile is not None:
            instance.system_id = profile.system_id
            instance.target_environment = profile.environment
        if not str(instance.title or "").strip():
            requirement = str(instance.requirement_text or "").strip()
            uploaded = self.cleaned_data.get("requirement_file")
            if requirement:
                instance.title = requirement.splitlines()[0][:80]
            elif uploaded:
                instance.title = Path(uploaded.name).stem[:80]
            else:  # guarded by clean(); retained for programmatic form use
                instance.title = "未命名测试任务"
        if commit:
            instance.save()
        return instance


class TestIntentImportForm(TestWorkflowForm):
    generation_count = forms.IntegerField(
        min_value=1,
        max_value=10,
        initial=1,
        label="生成数量",
        help_text="一次生成 1-10 个测试计划草稿，用于保留不同的生成候选。",
    )

    class Meta(TestWorkflowForm.Meta):
        model = TestIntentImport


class ApprovedTestPlanImportForm(forms.Form):
    """Upload one already-approved layer-one handoff without bypassing validation."""

    resource_profile = forms.ModelChoiceField(
        queryset=TestResourceProfile.objects.none(),
        label="测试资源",
        help_text="导入产物的被测系统和测试分类必须与该资源一致。",
    )
    artifact_file = forms.FileField(
        label="测试计划文件",
        help_text="请选择从可信测试设计流程导出的、已经完成审批的 JSON 文件。",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["resource_profile"].queryset = TestResourceProfile.objects.filter(
            enabled=True
        )

    def clean_artifact_file(self):
        uploaded = self.cleaned_data["artifact_file"]
        if int(getattr(uploaded, "size", 0) or 0) > 5 * 1024 * 1024:
            raise ValidationError("导入文件不能超过 5 MiB")
        raw = uploaded.read(5 * 1024 * 1024 + 1)
        if len(raw) > 5 * 1024 * 1024:
            raise ValidationError("导入文件不能超过 5 MiB")
        try:
            payload = json.loads(raw.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValidationError("导入文件必须是 UTF-8 JSON") from exc
        if not isinstance(payload, dict):
            raise ValidationError("导入文件顶层必须是 JSON 对象")
        return payload


class TestPlanExecutionInputForm(forms.ModelForm):
    runtime_variables = forms.CharField(
        required=False,
        label="执行数据（变量 / 值，可选）",
        widget=forms.Textarea(attrs={"rows": 6}),
        help_text=(
            "每行填写 名称=值，例如 account_id=A-1。UI 测试可能事先不知道页面会要求哪些输入；"
            "未注明的输入可能被模型自动编造，请在审批执行计划时重点检查。测试账号和密码可作为本次变量填写。"
        ),
    )
    class Meta:
        model = TestPlanArtifact
        fields = ("review_comments",)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        instance = getattr(self, "instance", None)
        if not instance or not instance.pk or self.is_bound:
            return
        previous = instance.execution_plans.order_by("-version", "-id").first()
        payload = dict(getattr(previous, "execution_input", None) or {})
        variables = payload.get("variables") or {}
        lines = []
        for name, value in variables.items():
            rendered = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False)
            lines.append(f"{name}={rendered}")
        self.initial["runtime_variables"] = "\n".join(lines)

    def clean_runtime_variables(self):
        raw = str(self.cleaned_data.get("runtime_variables") or "").strip()
        values = {}
        for line_number, line in enumerate(raw.splitlines(), start=1):
            text = line.strip()
            if not text:
                continue
            if "=" not in text:
                raise ValidationError(f"第 {line_number} 行必须使用 名称=值")
            name, raw_value = text.split("=", 1)
            name = name.strip()
            if not name:
                raise ValidationError(f"第 {line_number} 行变量名称不能为空")
            if name in values:
                raise ValidationError(f"变量 {name} 重复填写")
            raw_value = raw_value.strip()
            try:
                value = json.loads(raw_value)
            except json.JSONDecodeError:
                value = raw_value
            values[name] = value
        return values

    def execution_input(self):
        return {
            "schema_version": "test-runtime-input.v1",
            "variables": self.cleaned_data.get("runtime_variables") or {},
        }


class ExecutionRunDateFilterForm(forms.Form):
    date_from = forms.DateField(
        required=False,
        label="开始日期",
        widget=forms.DateInput(
            attrs={"type": "date", "aria-label": "选择开始日期"},
            format="%Y-%m-%d",
        ),
        input_formats=["%Y-%m-%d"],
    )
    date_to = forms.DateField(
        required=False,
        label="结束日期",
        widget=forms.DateInput(
            attrs={"type": "date", "aria-label": "选择结束日期"},
            format="%Y-%m-%d",
        ),
        input_formats=["%Y-%m-%d"],
    )

    def clean(self):
        cleaned = super().clean()
        date_from = cleaned.get("date_from")
        date_to = cleaned.get("date_to")
        if date_from and date_to and date_from > date_to:
            raise ValidationError("开始日期不能晚于结束日期")
        return cleaned


def _json(value):
    if hasattr(value, "model_dump"):
        return value.model_dump(mode="json")
    if isinstance(value, dict):
        return {key: _json(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json(item) for item in value]
    return value


def _business_text(value) -> str:
    if isinstance(value, dict):
        return str(value.get("text") or value.get("question") or "")
    return str(value or "")


def _display_lines(values):
    rendered = [str(value) for value in values if str(value).strip()]
    if not rendered:
        return "-"
    return format_html_join("", "{}<br>", ((value,) for value in rendered))


def _expected_line(item: dict) -> str:
    line = _business_text(item)
    if item.get("expected") is not None:
        expected = json.dumps(item["expected"], ensure_ascii=False, sort_keys=True)
        line += f" [{item.get('operator') or 'equals'} {expected}]"
    channel = item.get("channel_hint")
    return f"{TEST_CATEGORY_LABELS.get(channel, channel)}: {line}" if channel else line


def _assertion_line(assertion: dict) -> str:
    kind = assertion.get("kind") or "assert"
    subject = (
        assertion.get("path")
        or assertion.get("name")
        or assertion.get("column")
        or assertion.get("metric")
        or kind
    )
    operator = assertion.get("operator") or "equals"
    expected = assertion.get("expected")
    if expected is None and "value" in assertion:
        expected = assertion.get("value")
    rendered = json.dumps(expected, ensure_ascii=False, sort_keys=True)
    unit = assertion.get("unit") or ""
    return f"断言 {subject} {operator} {rendered}{unit}"


def _stage_resources(stage: dict) -> list[str]:
    execution = stage.get("execution") or {}
    kind = str(stage.get("executor_kind") or execution.get("kind") or "unknown")
    if kind in UI_EXECUTOR_KINDS:
        lines = [f"{stage.get('order', '?')}. UI 页面测试"]
        for row in execution.get("rows", []):
            lines.append(f"Action: {row.get('action') or row.get('operation_ref') or '-'}")
            for assertion in row.get("assertions", []):
                lines.append("Check: " + _assertion_line(assertion))
        return lines
    if kind == "http_api":
        lines = [f"{stage.get('order', '?')}. 接口测试"]
        for request in execution.get("requests", []):
            lines.append(
                f"请求: {request.get('method') or '-'} {request.get('path') or '-'}"
            )
            bindings = [
                f"{slot}={ref}"
                for item in request.get("data_bindings", [])
                for slot, ref in (item.get("input_refs") or {}).items()
            ]
            if bindings:
                lines.append("数据绑定: " + ", ".join(bindings))
            lines.extend(_assertion_line(item) for item in request.get("assertions", []))
        return lines
    if kind == "database":
        lines = [f"{stage.get('order', '?')}. 数据库测试（只读）"]
        for operation in execution.get("operations", []):
            lines.append(f"查询: {operation.get('operation_ref') or '-'}")
            lines.extend(_assertion_line(item) for item in operation.get("assertions", []))
        return lines
    if kind == "performance":
        lines = [
            f"{stage.get('order', '?')}. 性能/压力测试",
            f"负载配置: {execution.get('profile_ref') or '-'}",
            f"驱动: {execution.get('driver_ref') or '-'}",
        ]
        for load_stage in execution.get("stages", []):
            lines.append(
                "负载阶段: "
                + ", ".join(f"{key}={value}" for key, value in load_stage.items())
            )
        lines.extend(_assertion_line(item) for item in execution.get("thresholds", []))
        return lines
    if kind == "tcp_port":
        lines = [f"{stage.get('order', '?')}. TCP 端口测试"]
        for probe in execution.get("probes", []):
            lines.append(
                f"探测: {probe.get('host_ref') or '-'}:{probe.get('port') or '-'} "
                f"(超时 {probe.get('timeout_seconds') or '-'}s)"
            )
            lines.extend(_assertion_line(item) for item in probe.get("assertions", []))
        return lines
    return [f"{stage.get('order', '?')}. 未识别执行类型: {kind}"]


def _stage_has_operations(stage: dict) -> bool:
    execution = (stage or {}).get("execution") or {}
    kind = str(stage.get("executor_kind") or execution.get("kind") or "")
    if kind in UI_EXECUTOR_KINDS:
        return bool(execution.get("rows"))
    collection_by_kind = {
        "http_api": "requests",
        "database": "operations",
        "performance": "stages",
        "tcp_port": "probes",
    }
    collection = collection_by_kind.get(kind)
    return bool(collection and execution.get(collection))


def _generation_result_from_payload(value: dict):
    from .intent.contracts import (
        TestDesign,
        TestDesignCandidate,
        TestDesignInputSnapshot,
        TestDesignRequest,
        TestDesignValidationReport,
    )
    from .intent.service import TestDesignGenerationResult

    value = value or {}
    if not value:
        raise ValidationError("尚未生成测试计划")
    return TestDesignGenerationResult(
        request=TestDesignRequest.model_validate(value["request"]),
        candidate=TestDesignCandidate.model_validate(value["candidate"]),
        design=TestDesign.model_validate(value["design"]),
        input_snapshot=TestDesignInputSnapshot.model_validate(value["input_snapshot"]),
        validation=TestDesignValidationReport.model_validate(value["validation"]),
    )


def _compilation_result_from_payload(value: dict):
    from .planning.compiler import PlanCompilationResult
    from .planning.contracts import (
        ExecutorArtifactBundle,
        PlanValidationReport,
        TestPlanDraft,
    )

    value = value or {}
    if not value:
        raise ValidationError("尚未生成测试计划")
    return PlanCompilationResult(
        plan=TestPlanDraft.model_validate(value["plan"]),
        validation=PlanValidationReport.model_validate(value["validation"]),
        artifacts=[ExecutorArtifactBundle.model_validate(item) for item in value["artifacts"]],
    )


def _artifact_generation_lock(root: Path, workflow_id: str):
    """Prevent concurrent writers from cleaning or replacing the same plan tree."""

    lock_path = root / "workflows" / workflow_id / ".plan-generation.lock"
    return nonblocking_file_lock(
        lock_path,
        conflict_message="该任务正在生成执行计划，请等待当前生成完成",
    )


@admin.register(ApprovedKnowledgeEntry)
class ApprovedKnowledgeEntryAdmin(SingleRecordActionAdmin, admin.ModelAdmin):
    form = ApprovedKnowledgeEntryForm
    page_title = "业务知识库"
    page_description = "保存业务规则和历史经验；包括历史 SQL 及用途。知识只提供参考，实际执行仍受测试资源约束"
    list_display = ("title", "system_id", "approval_status", "updated_at")
    list_display_links = ("title",)
    list_filter = ("status", "system_id", "updated_at")
    search_fields = ("title", "content", "system_id")
    readonly_fields = (
        "status",
        "approval_summary",
        "created_at",
        "updated_at",
    )
    fieldsets = (
        (
            "知识内容",
            {
                "description": (
                    "格式：直接粘贴文字或上传常见文档，不要求固定模板。"
                    "作用：为模型提供可复用规则和历史经验；它不会扩大测试资源权限，也不会直接执行。"
                ),
                "fields": ("system_id", "title", "content", "source_file"),
            },
        ),
        ("发布状态", {"fields": ("status", "approval_summary", "updated_at")}),
    )
    actions = None

    def get_urls(self):
        custom = [
            path(
                "example/",
                self.admin_site.admin_view(self.download_example),
                name="test_platform_approvedknowledgeentry_example",
            )
        ]
        return custom + super().get_urls()

    def download_example(self, request):
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        return _example_download("business_knowledge.md", "text/markdown")

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            return (self.fieldsets[0],)
        return self.fieldsets

    def get_readonly_fields(self, request, obj=None):
        values = list(self.readonly_fields)
        if obj is not None and obj.status == ApprovedKnowledgeEntry.Status.APPROVED:
            values.extend(("system_id", "title", "content", "source_file"))
        return tuple(values)

    def get_record_commands(self, request, obj):
        if obj is None or obj.status == ApprovedKnowledgeEntry.Status.APPROVED:
            return []
        return [{
            "name": "_approve_knowledge",
            "label": "审核并发布",
            "handler": "run_approve_knowledge",
            "kind": "primary",
        }]

    def run_approve_knowledge(self, request, obj):
        try:
            obj.approve()
            self.message_user(
                request,
                f"“{obj.title}”已发布，现在可以在测试任务中选择。",
                messages.SUCCESS,
            )
        except ValidationError as exc:
            self.message_user(request, f"发布失败：{exc}", messages.ERROR)
        return HttpResponseRedirect(
            reverse("admin:test_platform_approvedknowledgeentry_change", args=[obj.pk])
        )

    @admin.display(description="状态")
    def approval_status(self, obj):
        return _status_badge(obj.status, obj.get_status_display())

    @admin.display(description="审批信息")
    def approval_summary(self, obj):
        if obj.status != ApprovedKnowledgeEntry.Status.APPROVED:
            return "保存后检查内容，确认无误再审核发布"
        return timezone.localtime(obj.approved_at).strftime("%Y-%m-%d %H:%M")


@admin.register(TestResourceProfile)
class TestResourceProfileAdmin(SingleRecordActionAdmin, admin.ModelAdmin):
    form = TestResourceProfileForm
    change_list_template = "admin/test_platform/resource_change_list.html"
    page_title = "测试资源"
    page_description = "一个系统和环境的一体化资源，可同时配置多种复合测试能力"

    class Media:
        js = ("test_platform/resource_form_v3.js",)

    list_display = (
        "name",
        "target_scope",
        "resource_channels",
        "enabled",
        "updated_at",
    )
    list_filter = ("enabled", "updated_at")
    search_fields = ("profile_id", "name")
    readonly_fields = (
        "profile_id",
        "resource_channels",
        "created_at",
        "updated_at",
    )
    _example_files = {
        "openapi": ("openapi.yaml", "application/yaml"),
        "database": ("database_queries.json", "application/json"),
        "performance": ("performance_profiles.json", "application/json"),
    }

    def get_urls(self):
        custom = [
            path(
                "examples/<slug:example_name>/",
                self.admin_site.admin_view(self.download_resource_example),
                name="test_platform_testresourceprofile_example",
            )
        ]
        return custom + super().get_urls()

    def download_resource_example(self, request, example_name):
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        definition = self._example_files.get(example_name)
        if definition is None:
            raise Http404("未知资源示例")
        filename, content_type = definition
        return _example_download(filename, content_type)

    def get_fieldsets(self, request, obj=None):
        example_name = "admin:test_platform_testresourceprofile_example"
        openapi_url = reverse(example_name, args=["openapi"])
        database_url = reverse(example_name, args=["database"])
        performance_url = reverse(example_name, args=["performance"])
        return (
            (
                "基本信息",
                {
                    "description": mark_safe(
                        '<div class="tb-resource-guide"><p><strong>格式：</strong>配置名称和被测系统。'
                        '<strong>作用：</strong>把同一系统的多种测试能力归在一起。</p></div>'
                    ),
                    "fields": (
                        "name",
                        "system_id",
                        "resource_types",
                        "enabled",
                    ),
                },
            ),
            (
                "UI 页面测试资源",
                {
                    "classes": ("resource-section", "resource-ui"),
                    "description": mark_safe(
                        '<div class="tb-resource-guide"><p><strong>网页 Agent：</strong>'
                        '文件或文字只包含 URL、页面/功能和最大步数。'
                        '例如：<code>https://shop.example.test | 商品购物车 | 20</code>。'
                        '不要填写控件步骤、所需变量、账号密码或执行历史。</p></div>'
                    ),
                    "fields": (
                        "ui_agent_asset_file",
                        "ui_agent_asset_text",
                    ),
                },
            ),
            (
                "接口测试资源",
                {
                    "classes": ("resource-section", "resource-api"),
                    "description": format_html(
                        '<div class="tb-resource-guide"><p><strong>格式：</strong>OpenAPI 或大致描述 method、path、参数的文档/文字。'
                        '<strong>作用：</strong>让模型整理当前环境可调用的接口目录；基础地址仍精确填写。</p>'
                        '<a class="tb-example-download" href="{}">下载可选 OpenAPI 示例</a></div>',
                        openapi_url,
                    ),
                    "fields": ("api_openapi_file", "api_asset_text", "api_base_url"),
                },
            ),
            (
                "数据库测试资源",
                {
                    "classes": ("resource-section", "resource-database"),
                    "description": format_html(
                        '<div class="tb-resource-guide"><p><strong>格式：</strong>DDL、数据字典、表结构文档或简单文字。'
                        '<strong>作用：</strong>让模型整理可读表字段，并在该范围内生成只读 SQL 草稿。</p>'
                        '<a class="tb-example-download" href="{}">下载可选结构化示例</a></div>',
                        database_url,
                    ),
                    "fields": (
                        "database_query_file",
                        "database_asset_text",
                        "database_connection_ref",
                    ),
                },
            ),
            (
                "性能/压力测试资源",
                {
                    "classes": ("resource-section", "resource-performance"),
                    "description": format_html(
                        '<div class="tb-resource-guide"><p><strong>格式：</strong>性能要求、历史方案、现有配置或简单文字。'
                        '<strong>作用：</strong>让模型整理目标、负载安全上限和可观测指标。</p>'
                        '<a class="tb-example-download" href="{}">下载可选结构化示例</a></div>',
                        performance_url,
                    ),
                    "fields": ("performance_profile_file", "performance_asset_text"),
                },
            ),
            (
                "TCP 端口测试资源",
                {
                    "classes": ("resource-section", "resource-port"),
                    "description": mark_safe(
                        '<div class="tb-resource-guide"><p><strong>格式：</strong>一个主机名/IP 和一个端口号。'
                        '<strong>作用：</strong>验证指定端点的 TCP 连通性和连接耗时，不扫描端口范围。</p></div>'
                    ),
                    "fields": ("port_host", "port_number"),
                },
            ),
        )

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        category = request.GET.get("category")
        if category in TEST_CATEGORY_LABELS:
            initial["resource_types"] = [category]
        return initial

    @admin.display(description="已配置渠道")
    def resource_channels(self, obj):
        labels = {
            "ui": "UI",
            "api": "API",
            "database": "数据库",
            "performance": "性能",
            "port": "端口",
        }
        return "、".join(labels[item] for item in sorted(obj.configured_channels())) or "未配置"

    @admin.display(description="被测目标")
    def target_scope(self, obj):
        return obj.system_id or "-"


class TestWorkflowAdmin(SingleRecordActionAdmin, admin.ModelAdmin):
    form = TestWorkflowForm
    polling_statuses = (TestWorkflow.Status.DESIGN_GENERATING,)

    list_display = (
        "workflow_id",
        "title",
        "status",
        "resource_profile",
        "updated_at",
    )
    list_filter = ("status", "target_environment", "updated_at")
    search_fields = ("workflow_id", "title", "system_id")
    readonly_fields = (
        "workflow_id",
        "status",
        "generation_progress_display",
        "last_error",
        "created_at",
        "updated_at",
    )
    actions = ("mark_selected", "unmark_selected", "delete_selected")

    def get_form(self, request, obj=None, change=False, **kwargs):
        form_class = super().get_form(request, obj, change, **kwargs)
        resource_field = form_class.base_fields.get("resource_profile")
        widget = getattr(resource_field, "widget", None)
        for attribute in (
            "can_add_related",
            "can_change_related",
            "can_delete_related",
            "can_view_related",
        ):
            if widget is not None and hasattr(widget, attribute):
                setattr(widget, attribute, False)
        return form_class
    def get_fieldsets(self, request, obj=None):
        sections = [
            (
                "测试输入",
                {
                    "description": (
                        "格式：自然语言或常见需求文件。作用：说明本次要测什么、期望什么；"
                        "模型会生成测试计划草稿供人工审批。"
                    ),
                    "fields": (
                        "resource_profile",
                        "requirement_text",
                        "requirement_file",
                        "allowed_channels",
                    )
                },
            ),
            (
                "知识增强（可选）",
                {
                    "description": "选择经过审核的业务规则作为设计参考；系统不会用它覆盖原始需求。",
                    "fields": ("knowledge_scope_ids",),
                },
            ),
        ]
        if obj is not None:
            sections.append(
                (
                    "当前结果",
                    {
                        "fields": (
                            "status",
                            "generation_progress_display",
                            "last_error",
                            "updated_at",
                        )
                    },
                )
            )
        return sections

    @admin.display(description="生成进度")
    def generation_progress_display(self, obj):
        return _generation_progress_display(obj)

    @admin.display(description="待审核测试计划")
    def design_review_preview(self, obj):
        design = (obj.design_generation or {}).get("design") or (
            (obj.approved_design_bundle or {}).get("design") or {}
        )
        if not design:
            return "尚未生成测试计划"
        scenario_sections = []
        for scenario in design.get("scenarios", []):
            categories = sorted(
                {
                    item.get("channel_hint")
                    for item in [
                        *scenario.get("operations", []),
                        *scenario.get("expected_results", []),
                    ]
                    if item.get("channel_hint")
                }
            )
            scenario_operations = scenario.get("operations", [])
            operation_positions = {
                item.get("operation_id"): index
                for index, item in enumerate(scenario_operations, start=1)
                if item.get("operation_id")
            }
            expected_by_position = {}
            unmatched_expected = []
            for item in scenario.get("expected_results", []):
                position = operation_positions.get(item.get("after_operation_id"))
                if position is None:
                    position = item.get("after_operation_index")
                if position:
                    expected_by_position.setdefault(position, []).append(
                        _expected_line(item)
                    )
                else:
                    unmatched_expected.append(_expected_line(item))
            operation_rows = []
            for index, item in enumerate(scenario_operations, start=1):
                checks = expected_by_position.get(index, [])
                check_html = (
                    format_html(
                        "<ul class='tb-plan-checks'>{}</ul>",
                        format_html_join("", "<li>{}</li>", ((value,) for value in checks)),
                    )
                    if checks
                    else ""
                )
                operation_rows.append(
                    format_html(
                        "<li class='tb-plan-step'><span>{}</span><div><strong>{}</strong>"
                        "<p>{}</p>{}</div></li>",
                        index,
                        TEST_CATEGORY_LABELS.get(
                            item.get("channel_hint"), item.get("channel_hint") or "测试操作"
                        ),
                        _business_text(item),
                        check_html,
                    )
                )
            data_and_states = [
                "前置状态：" + _business_text(item)
                for item in scenario.get("required_states", [])
            ]
            for item in scenario.get("data_requirements", []):
                constraints = [
                    _business_text(value) for value in item.get("constraints", [])
                ]
                line = "测试数据：" + _business_text(item)
                if constraints:
                    line += "；约束：" + "；".join(constraints)
                data_and_states.append(line)
            impact = (scenario.get("state_impact") or {}).get("impact") or "-"
            metadata = format_html(
                "<div class='tb-plan-meta'><span><strong>测试分类：</strong>{}</span>"
                "<span><strong>覆盖方式：</strong>{}</span>"
                "<span><strong>状态影响：</strong>{}</span></div>",
                _category_badges(categories),
                "、".join(
                    TECHNIQUE_LABELS.get(str(item), str(item))
                    for item in scenario.get("techniques", [])
                ) or "未指定",
                impact,
            )
            supplemental = []
            if data_and_states:
                supplemental.append(
                    format_html(
                        "<div class='tb-plan-notes'><strong>前置与测试数据</strong>{}</div>",
                        format_html_join("", "<p>{}</p>", ((value,) for value in data_and_states)),
                    )
                )
            if unmatched_expected:
                supplemental.append(
                    format_html(
                        "<div class='tb-plan-notes'><strong>场景级检查</strong>{}</div>",
                        format_html_join("", "<p>{}</p>", ((value,) for value in unmatched_expected)),
                    )
                )
            scenario_sections.append(
                format_html(
                    "<section class='tb-plan-scenario'><header><h3>{}</h3>{}</header>"
                    "<ol class='tb-plan-steps'>{}</ol>{}</section>",
                    scenario.get("title") or "-",
                    metadata,
                    format_html_join("", "{}", ((row,) for row in operation_rows)),
                    format_html_join("", "{}", ((item,) for item in supplemental)),
                )
            )
        questions = [
            _business_text(item) for item in design.get("open_questions", [])
        ]
        categories = _design_categories(design)
        questions_html = (
            format_html(
                "<section class='tb-plan-questions'><strong>待确认问题</strong>{}</section>",
                format_html_join("", "<p>{}</p>", ((value,) for value in questions)),
            )
            if questions
            else ""
        )
        return format_html(
            "<div class='tb-review-summary tb-plan-review'><header class='tb-plan-overview'>"
            "<div><strong>{}</strong><p>{}</p></div>"
            "<span><strong>测试分类：</strong>{}</span></header>{}{}</div>",
            design.get("title") or "-",
            _business_text(design.get("objective")),
            _category_badges(categories),
            format_html_join("", "{}", ((item,) for item in scenario_sections)),
            questions_html,
        )

    @admin.display(description="待审核执行计划")
    def plan_review_preview(self, obj):
        plan = (obj.plan_compilation or {}).get("plan") or {}
        if not plan:
            return "尚未生成执行计划"
        flow_sections = []
        for flow in plan.get("flows", []):
            visible_stages = [
                stage for stage in flow.get("stages", []) if _stage_has_operations(stage)
            ]
            if not visible_stages:
                continue
            flow_categories = list(
                dict.fromkeys(
                    EXECUTOR_CATEGORY.get(
                        str(stage.get("executor_kind") or ""),
                        str(stage.get("executor_kind") or ""),
                    )
                    for stage in visible_stages
                    if stage.get("executor_kind")
                )
            )
            stage_rows = []
            for stage in sorted(
                visible_stages,
                key=lambda value: (value.get("order") or 0, value.get("stage_id") or ""),
            ):
                executor = str(
                    stage.get("executor_kind")
                    or (stage.get("execution") or {}).get("kind")
                    or ""
                )
                category = EXECUTOR_CATEGORY.get(executor, executor)
                details = _stage_resources(stage)[1:]
                stage_rows.append(
                    format_html(
                        "<li class='tb-flow-stage'><span>{}</span><div><strong>{}</strong>{}</div></li>",
                        stage.get("order") or "?",
                        TEST_CATEGORY_LABELS.get(category, category or "测试操作"),
                        format_html_join("", "<p>{}</p>", ((line,) for line in details)),
                    )
                )
            cleanup = flow.get("cleanup") or {}
            cleanup_html = (
                format_html("<p class='tb-flow-cleanup'><strong>完成后：</strong>清理测试数据</p>")
                if cleanup
                else ""
            )
            flow_sections.append(
                format_html(
                    "<section class='tb-plan-scenario tb-execution-flow'><header><h3>{}</h3>{}</header>"
                    "<ol class='tb-plan-steps'>{}</ol>{}</section>",
                    flow.get("name") or "-",
                    _category_badges(flow_categories),
                    format_html_join("", "{}", ((row,) for row in stage_rows)),
                    cleanup_html,
                )
            )
        questions = [
            _business_text(item) for item in plan.get("open_questions", [])
        ]
        plan_categories = _plan_categories(plan)
        stage_counts = _category_stage_counts(plan)
        count_text = "、".join(
            f"{TEST_CATEGORY_LABELS.get(category, category)} {count} 个执行阶段"
            for category, count in stage_counts.items()
        ) or "尚未生成执行阶段"
        questions_html = (
            format_html(
                "<section class='tb-plan-questions'><strong>待确认问题</strong>{}</section>",
                format_html_join("", "<p>{}</p>", ((value,) for value in questions)),
            )
            if questions
            else ""
        )
        return format_html(
            "<div class='tb-review-summary tb-plan-review'><header class='tb-plan-overview'>"
            "<div><strong>执行目标：{}</strong><p>{}</p></div>"
            "<span><strong>测试分类：</strong>{}</span></header>{}{}</div>",
            plan.get("target_system_id") or "-",
            count_text,
            _category_badges(plan_categories),
            format_html_join("", "{}", ((item,) for item in flow_sections)),
            questions_html,
        )

    @admin.action(description="生成测试计划")
    def generate_design(self, request, queryset, *, count=1):
        for obj in queryset:
            try:
                queue_design_generation(obj, count=count)
                self.message_user(
                    request,
                    f"“{obj.title}”已提交生成 {count} 个测试计划，完成后会出现在测试计划审批列表。",
                    messages.INFO,
                )
            except Exception as exc:
                self.message_user(request, f"“{obj.title}”生成失败：{exc}", messages.ERROR)

@admin.register(TestIntentImport)
class TestIntentImportAdmin(TestWorkflowAdmin):
    """Submit the human requirement and choose executable test categories."""

    form = TestIntentImportForm
    change_list_template = "admin/test_platform/intent_change_list.html"
    page_title = "测试意图"
    page_description = "提交需求并选择要生成的测试分类"
    list_display = (
        "marked_status",
        "task_title",
        "category_display",
        "workflow_status",
        "resource_profile",
        "updated_at",
    )
    list_display_links = ("task_title",)
    list_filter = (MarkedRecordFilter, TestCategoryFilter)
    search_fields = ("workflow_id", "title", "requirement_text")
    actions = ("mark_selected", "unmark_selected", "delete_selected")

    def get_urls(self):
        custom = [
            path(
                "example/",
                self.admin_site.admin_view(self.download_requirement_example),
                name="test_platform_testintentimport_example",
            )
        ]
        return custom + super().get_urls()

    def download_requirement_example(self, request):
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        return _example_download("test_requirement.md", "text/markdown")

    @admin.display(description="测试分类")
    def category_display(self, obj):
        return _category_badges(obj.allowed_channels)

    @admin.display(description="测试任务", ordering="title")
    def task_title(self, obj):
        return obj.title

    @admin.display(description="当前阶段", ordering="status")
    def workflow_status(self, obj):
        badge = _status_badge(obj.status, obj.get_status_display())
        if obj.status != TestWorkflow.Status.DESIGN_GENERATING:
            return badge
        return format_html(
            "{}{}",
            badge,
            _generation_progress_display(obj, compact=True),
        )

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        category = request.GET.get("category")
        if category in TEST_CATEGORY_LABELS:
            initial["allowed_channels"] = [category]
        return initial

    def get_fieldsets(self, request, obj=None):
        fieldsets = list(super().get_fieldsets(request, obj))
        title, options = fieldsets[0]
        fields = list(options["fields"])
        fields.append("generation_count")
        fieldsets[0] = (title, {**options, "fields": tuple(fields)})
        return tuple(fieldsets)

    def get_record_commands(self, request, obj):
        if obj and obj.status != TestWorkflow.Status.DESIGN_GENERATING:
            return [
                {
                    "name": "_generate_test_plan",
                    "label": "生成测试计划",
                    "handler": "run_generate_test_plan",
                    "kind": "primary",
                }
            ]
        return []

    def run_generate_test_plan(self, request, obj):
        form = self.get_form(request, obj)(request.POST, request.FILES, instance=obj)
        if not form.is_valid():
            self.message_user(request, "生成数量或测试输入无效。", messages.ERROR)
            return HttpResponseRedirect(request.path)
        self.generate_design(
            request,
            self.model.objects.filter(pk=obj.pk),
            count=form.cleaned_data["generation_count"],
        )
        return HttpResponseRedirect(reverse("admin:test_platform_testintentimport_changelist"))

    def show_record_save(self, request, obj) -> bool:
        return bool(
            obj is None or obj.status != TestWorkflow.Status.DESIGN_GENERATING
        )

    def get_queryset(self, request):
        return super().get_queryset(request)


def _artifact_preview_proxy(*, design=None, plan=None):
    class PreviewProxy:
        design_generation = {"design": design} if design else {}
        approved_design_bundle = {}
        plan_compilation = {"plan": plan} if plan else {}

    return PreviewProxy()


@admin.register(TestPlanArtifact)
class TestPlanArtifactAdmin(
    AuditPayloadAdminMixin,
    ApprovalStateTabsMixin,
    SingleRecordActionAdmin,
    admin.ModelAdmin,
):
    form = TestPlanExecutionInputForm
    change_list_template = "admin/test_platform/approval_change_list.html"
    page_title = "测试计划审批"
    page_description = "业务场景、覆盖方式与预期结果"
    audit_payload_specs = (
        (
            "generation_result",
            "生成记录",
            "退回修订时恢复模型候选、输入快照和规则校验结果",
        ),
        (
            "review_payload",
            "审批记录",
            "绑定审核决定、意见、审核人和内容指纹",
        ),
        (
            "approved_bundle",
            "下层交接包",
            "审批通过后作为生成执行计划的唯一正式输入",
        ),
    )
    list_display = (
        "marked_status",
        "title",
        "category_display",
        "approval_status",
        "source_kind",
        "updated_at",
    )
    list_display_links = ("title",)
    list_filter = (MarkedRecordFilter, ApprovalStateParameterFilter, TestCategoryFilter)
    search_fields = ("artifact_id", "design_id", "title")
    readonly_fields = (
        "artifact_id",
        "source_intent",
        "source_kind",
        "resource_profile",
        "title",
        "test_categories",
        "category_summary",
        "design_id",
        "version",
        "content_hash",
        "test_plan_preview",
        "knowledge_usage_summary",
        "system_audit_summary",
        "status",
        "last_error",
        "generation_result",
        "review_payload",
        "approved_bundle",
        "created_at",
        "updated_at",
    )
    fieldsets = (
        ("测试计划", {"fields": ("test_plan_preview",)}),
        ("设计参考", {"fields": ("knowledge_usage_summary",)}),
        (
            "计划来源",
            {
                "fields": (
                    "source_intent",
                    "source_kind",
                    "resource_profile",
                    "category_summary",
                )
            },
        ),
        ("当前结果", {"fields": ("status", "last_error", "updated_at")}),
        (
            "审批处理",
            {
                "description": "填写修改要求后，可按当前计划重新生成新的测试计划版本。",
                "fields": ("review_comments",),
            },
        ),
    )
    actions = ("mark_selected", "unmark_selected", "delete_selected")
    approval_state_statuses = {
        "pending": (TestPlanArtifact.Status.REVIEW,),
        "completed": (TestPlanArtifact.Status.APPROVED,),
        "failed": (
            TestPlanArtifact.Status.BLOCKED,
            TestPlanArtifact.Status.CHANGES,
            TestPlanArtifact.Status.ERROR,
        ),
    }

    def get_urls(self):
        custom = [
            path(
                "import/example/",
                self.admin_site.admin_view(self.download_import_example),
                name="test_platform_testplanartifact_import_example",
            ),
            path(
                "import/",
                self.admin_site.admin_view(self.import_approved_artifact),
                name="test_platform_testplanartifact_import",
            )
        ]
        return custom + super().get_urls()

    def download_import_example(self, request):
        return _example_download("approved_test_plan.json", "application/json")

    def import_approved_artifact(self, request):
        form = ApprovedTestPlanImportForm(
            request.POST or None,
            request.FILES or None,
        )
        if request.method == "POST" and form.is_valid():
            from .artifact_store import import_approved_test_plan

            try:
                artifact = import_approved_test_plan(
                    form.cleaned_data["artifact_file"],
                    form.cleaned_data["resource_profile"],
                )
            except Exception as exc:
                form.add_error(None, str(exc))
            else:
                self.message_user(
                    request,
                    f"“{artifact.title}”已导入并完成校验。",
                    messages.SUCCESS,
                )
                return HttpResponseRedirect(
                    reverse(
                        "admin:test_platform_testplanartifact_change",
                        args=[artifact.pk],
                    )
                )
        context = {
            **self.admin_site.each_context(request),
            "title": "导入已审批测试计划",
            "opts": self.model._meta,
            "form": form,
        }
        return TemplateResponse(
            request,
            "admin/test_platform/test_plan_import.html",
            context,
        )

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return True

    @staticmethod
    def _can_generate_execution_plan(obj):
        return obj.status in {
            TestPlanArtifact.Status.REVIEW,
            TestPlanArtifact.Status.APPROVED,
        }

    def get_fieldsets(self, request, obj=None):
        fieldsets = list(super().get_fieldsets(request, obj))
        if obj is not None and self._can_generate_execution_plan(obj):
            fieldsets.insert(
                -1,
                (
                    "执行计划输入",
                    {
                        "description": (
                            "这些本次运行变量将冻结到即将生成的执行计划中；"
                            "执行和同计划重新运行都会复用同一份输入。"
                        ),
                        "fields": ("runtime_variables",),
                    },
                ),
            )
        return tuple(fieldsets)

    @admin.display(description="测试分类")
    def category_display(self, obj):
        return _category_badges(obj.test_categories)

    @admin.display(description="测试分类")
    def category_summary(self, obj):
        return _category_badges(obj.test_categories)

    @admin.display(description="本次采用的业务知识")
    def knowledge_usage_summary(self, obj):
        snapshot = (obj.generation_result or {}).get("input_snapshot") or {}
        documents = snapshot.get("approved_knowledge") or []
        if not documents:
            return mark_safe(
                "<div class='tb-knowledge-empty'><strong>未使用额外业务知识</strong>"
                "<span>本计划依据需求原文、测试分类和覆盖方式生成。</span></div>"
            )
        cards = []
        for document in documents:
            content = str(document.get("content") or "").strip()
            cards.append(
                format_html(
                    "<article class='tb-knowledge-card'><strong>已审核业务规则</strong>"
                    "<p>{}</p></article>",
                    content,
                )
            )
        return format_html(
            "<div class='tb-knowledge-summary'><p>本次测试设计参考了 {} 条已审核业务规则。</p>{}</div>",
            len(cards),
            format_html_join("", "{}", ((card,) for card in cards)),
        )

    @admin.display(description="审批状态", ordering="status")
    def approval_status(self, obj):
        return _status_badge(obj.status, obj.get_status_display())

    def get_record_commands(self, request, obj):
        if obj is None:
            return []
        commands = []
        can_generate_execution = self._can_generate_execution_plan(obj)
        if can_generate_execution:
            commands.append(
                {
                    "name": "_approve_test_plan",
                    "label": "通过并生成执行计划" if obj.status == TestPlanArtifact.Status.REVIEW else "生成执行计划",
                    "handler": "run_approve_test_plan",
                    "kind": "primary",
                }
            )
        if obj.status in {
            TestPlanArtifact.Status.REVIEW,
            TestPlanArtifact.Status.BLOCKED,
            TestPlanArtifact.Status.CHANGES,
            TestPlanArtifact.Status.APPROVED,
        }:
            if obj.source_intent_id:
                commands.append(
                    {
                        "name": "_revise_test_plan",
                        "label": "按修改需求重新生成当前计划",
                        "handler": "run_revise_test_plan",
                        "kind": "secondary",
                    }
                )
        return commands

    def show_record_save(self, request, obj) -> bool:
        return bool(
            obj
            and obj.status
            in {
                TestPlanArtifact.Status.REVIEW,
                TestPlanArtifact.Status.BLOCKED,
                TestPlanArtifact.Status.CHANGES,
                TestPlanArtifact.Status.APPROVED,
            }
        )

    def get_readonly_fields(self, request, obj=None):
        fields = list(super().get_readonly_fields(request, obj))
        if obj and not self.show_record_save(request, obj):
            fields.append("review_comments")
        return tuple(dict.fromkeys(fields))

    def run_approve_test_plan(self, request, obj):
        form = self.get_form(request, obj)(request.POST, instance=obj)
        if not form.is_valid():
            self.message_user(
                request,
                "；".join(
                    str(message)
                    for field_errors in form.errors.values()
                    for message in field_errors
                ),
                messages.ERROR,
            )
            return HttpResponseRedirect(
                reverse("admin:test_platform_testplanartifact_change", args=[obj.pk])
            )
        self.approve_and_generate_execution_plan(
            request,
            self.model.objects.filter(pk=obj.pk),
            execution_input=form.execution_input(),
        )
        return HttpResponseRedirect(
            reverse("admin:test_platform_executionplanartifact_changelist")
        )

    def run_revise_test_plan(self, request, obj):
        comments = str(obj.review_comments or "").strip()
        if not comments:
            self.message_user(request, "请先填写修改要求，再重新生成测试计划。", messages.ERROR)
            return HttpResponseRedirect(
                reverse("admin:test_platform_testplanartifact_change", args=[obj.pk])
            )
        if not obj.source_intent_id:
            self.message_user(request, "导入的已审批产物不能在本步骤直接修订。", messages.ERROR)
            return HttpResponseRedirect(
                reverse("admin:test_platform_testplanartifact_change", args=[obj.pk])
            )

        obj.status = TestPlanArtifact.Status.CHANGES
        obj.save(update_fields=("status", "updated_at"))
        queue_design_generation(
            obj.source_intent,
            previous_artifact_id=obj.pk,
        )
        self.message_user(request, "已按修改要求重新生成测试计划。", messages.INFO)
        return HttpResponseRedirect(reverse("admin:test_platform_testplanartifact_changelist"))

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("source_intent")

    @admin.display(description="待审批测试计划")
    def test_plan_preview(self, obj):
        design = (obj.generation_result or {}).get("design") or (
            (obj.approved_bundle or {}).get("design") or {}
        )
        preview = TestWorkflowAdmin.design_review_preview(
            self,
            _artifact_preview_proxy(design=design),
        )
        validation = (obj.generation_result or {}).get("validation") or {}
        if validation.get("passed") is True:
            return preview
        findings = validation.get("findings") or []
        if not findings:
            return preview
        return format_html(
            "{}<div class='errornote'><strong>规则校验未通过：</strong>{}</div>",
            preview,
            _display_lines(item.get("message") for item in findings if item.get("blocking")),
        )

    @admin.action(description="审批测试计划 / 生成执行计划")
    def approve_and_generate_execution_plan(
        self,
        request,
        queryset,
        *,
        execution_input=None,
    ):
        from .intent.contracts import ReviewDecision
        from .approval_service import persist_test_plan_approval
        from .service_factory import get_workflow

        service = get_workflow()
        for artifact in queryset:
            try:
                if artifact.status == TestPlanArtifact.Status.REVIEW:
                    output = service.review_design(
                        _generation_result_from_payload(artifact.generation_result),
                        decision=ReviewDecision.APPROVED,
                        comments=artifact.review_comments or "后台审核通过",
                    )
                    artifact = persist_test_plan_approval(
                        artifact.pk,
                        expected_status=artifact.status,
                        review_payload=_json(output.review),
                        approved_bundle=_json(output.approved_bundle),
                    )
                artifact.refresh_from_db()
                feedback = str(artifact.review_comments or "").strip() or None
                execution_artifact = queue_execution_plan_generation(
                    artifact,
                    feedback=feedback,
                    execution_input=execution_input,
                )
                artifact.last_error = ""
                artifact.save(update_fields=("last_error", "updated_at"))
                self.message_user(
                    request,
                    f"“{execution_artifact.title}”已提交生成。",
                    messages.INFO,
                )
            except Exception as exc:
                # A stale/concurrent approval is a request conflict, not an
                # artifact processing error; never overwrite the winning audit.
                if not isinstance(exc, ValidationError):
                    TestPlanArtifact.objects.filter(pk=artifact.pk).update(
                        last_error=str(exc)[:20_000]
                    )
                self.message_user(
                    request,
                    f"“{artifact.title}”处理失败：{exc}",
                    messages.ERROR,
                )

@admin.register(ExecutionPlanArtifact)
class ExecutionPlanArtifactAdmin(
    AuditPayloadAdminMixin,
    SingleRecordActionAdmin,
    admin.ModelAdmin,
):
    change_list_template = "admin/test_platform/change_list_base.html"
    page_title = "执行计划"
    page_description = "审批后可单个或批量运行，同一计划可以重复运行"
    polling_statuses = (ExecutionPlanArtifact.Status.GENERATING,)
    audit_payload_specs = (
        (
            "catalog_snapshot",
            "资源目录快照",
            "执行前比较当前资源，资源发生漂移时阻止旧计划运行",
        ),
        (
            "compilation_result",
            "编译记录",
            "审批、修订并定位每个 flow、stage 和执行文件",
        ),
        (
            "review_payload",
            "审批记录",
            "绑定审批决定、意见、审核人和计划指纹",
        ),
        (
            "approved_bundle",
            "执行交接包",
            "审批通过后作为执行协调器唯一接受的正式输入",
        ),
    )
    list_display = (
        "marked_status",
        "title",
        "source_kind_display",
        "category_display",
        "execution_status",
        "updated_at",
    )
    list_display_links = ("title",)
    list_filter = (
        MarkedRecordFilter,
        ExecutionPlanStateFilter,
        ExecutionPlanSourceFilter,
        TestCategoryFilter,
    )
    search_fields = (
        "artifact_id",
        "plan_id",
        "title",
        "source_test_plan__artifact_id",
        "source_test_plan__design_id",
        "source_test_plan__title",
        "source_test_plan__source_intent__workflow_id",
        "source_test_plan__source_intent__title",
    )
    readonly_fields = (
        "artifact_id",
        "source_test_plan",
        "source_kind_display",
        "plan_source_summary",
        "resource_profile",
        "resource_profile_summary",
        "title",
        "test_categories",
        "category_summary",
        "plan_id",
        "version",
        "content_hash",
        "execution_plan_preview",
        "execution_input_summary",
        "approval_focus_preview",
        "page_execution_basis",
        "validation_summary",
        "execution_artifact_preview",
        "system_audit_summary",
        "status",
        "generation_progress_display",
        "last_error",
        "catalog_snapshot",
        "compilation_result",
        "review_payload",
        "approved_bundle",
        "artifact_root_ref",
        "created_at",
        "updated_at",
    )
    fieldsets = (
        (
            "执行计划",
            {
                "fields": (
                    "execution_plan_preview",
                    "execution_input_summary",
                    "approval_focus_preview",
                    "validation_summary",
                    "page_execution_basis",
                    "execution_artifact_preview",
                )
            },
        ),
        (
            "计划来源",
            {
                "fields": (
                    "plan_source_summary",
                    "resource_profile_summary",
                    "category_summary",
                )
            },
        ),
        (
            "当前结果",
            {
                "fields": (
                    "status",
                    "generation_progress_display",
                    "last_error",
                    "updated_at",
                )
            },
        ),
        (
            "审批处理",
            {
                "description": "填写修改要求后，可基于当前测试计划重新生成执行计划。",
                "fields": ("review_comments",),
            },
        ),
    )
    actions = (
        "run_selected",
        "mark_selected",
        "unmark_selected",
        "delete_selected",
    )

    def get_urls(self):
        opts = self.model._meta
        custom = [
            path(
                "<path:object_id>/artifact/<str:flow_id>/<str:stage_id>/<path:path_ref>",
                self.admin_site.admin_view(self.download_execution_artifact),
                name=f"{opts.app_label}_{opts.model_name}_execution_artifact",
            )
        ]
        return custom + super().get_urls()

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return True

    def get_fieldsets(self, request, obj=None):
        fieldsets = list(super().get_fieldsets(request, obj))
        if obj is None:
            return tuple(fieldsets)
        plan = (obj.compilation_result or {}).get("plan") or (
            (obj.approved_bundle or {}).get("plan") or {}
        )
        stages = [
            stage
            for flow in plan.get("flows") or []
            for stage in flow.get("stages") or []
            if _stage_has_operations(stage)
        ]
        has_ui = any(
            str(
                stage.get("executor_kind")
                or (stage.get("execution") or {}).get("kind")
                or ""
            )
            in UI_EXECUTOR_KINDS
            for stage in stages
        )
        execution_fields = ["execution_plan_preview"]
        if (obj.execution_input or {}).get("variables"):
            execution_fields.append("execution_input_summary")
        if stages:
            execution_fields.append("approval_focus_preview")
        execution_fields.append("validation_summary")
        if has_ui:
            execution_fields.append("page_execution_basis")
        if any(
            artifact.get("artifact_refs")
            for artifact in (obj.compilation_result or {}).get("artifacts") or []
        ):
            execution_fields.append("execution_artifact_preview")
        title, options = fieldsets[0]
        fieldsets[0] = (title, {**options, "fields": tuple(execution_fields)})
        return tuple(fieldsets)

    @admin.display(description="测试分类")
    def category_display(self, obj):
        return _category_badges(obj.test_categories)

    @admin.display(description="计划来源", ordering="source_test_plan__source_kind")
    def source_kind_display(self, obj):
        source_kind = obj.source_test_plan.source_kind
        label = (
            "手动导入测试计划"
            if source_kind == TestPlanArtifact.SourceKind.IMPORTED
            else "测试意图生成"
        )
        return format_html(
            '<span class="tb-source-kind tb-source-kind--{}">{}</span>',
            source_kind,
            label,
        )

    @admin.display(description="来源与业务知识")
    def plan_source_summary(self, obj):
        source_plan = obj.source_test_plan
        plan_url = reverse(
            "admin:test_platform_testplanartifact_change",
            args=[source_plan.pk],
        )
        if source_plan.source_kind == TestPlanArtifact.SourceKind.IMPORTED:
            source_label = "手动导入已审批测试计划"
            intent = format_html(
                "<span class='tb-source-note'>无关联测试意图；导入文件直接提供测试计划。</span>"
            )
            empty_knowledge = "导入文件未包含业务知识库使用记录"
        else:
            source_label = "根据测试意图生成"
            if source_plan.source_intent_id:
                source_intent = source_plan.source_intent
                intent = format_html(
                    "<a href='{}'>{}</a>",
                    reverse(
                        "admin:test_platform_testintentimport_change",
                        args=[source_intent.pk],
                    ),
                    source_intent.title or source_intent.workflow_id,
                )
            else:
                intent = format_html(
                    "<span class='tb-capability-warning'>缺少来源测试意图</span>"
                )
            empty_knowledge = "本计划未使用业务知识库"

        snapshot = (source_plan.generation_result or {}).get("input_snapshot") or {}
        documents = snapshot.get("approved_knowledge") or []
        if documents:
            knowledge = format_html(
                "<ul>{}</ul>",
                format_html_join(
                    "",
                    "<li>{}</li>",
                    (
                        (str(item.get("content") or "已审批业务知识").strip(),)
                        for item in documents
                    ),
                ),
            )
        else:
            knowledge = format_html(
                "<span class='tb-source-note'>{}</span>",
                empty_knowledge,
            )
        return format_html(
            "<div class='tb-plan-source-summary'>"
            "<dl><div><dt>生成方式</dt><dd>{}</dd></div>"
            "<div><dt>来源测试计划</dt><dd><a href='{}'>{}</a></dd></div>"
            "<div><dt>来源测试意图</dt><dd>{}</dd></div></dl>"
            "<div class='tb-plan-knowledge'><strong>本次采用的业务知识</strong>{}</div>"
            "</div>",
            source_label,
            plan_url,
            source_plan.title or source_plan.artifact_id,
            intent,
            knowledge,
        )

    @admin.display(description="测试资源")
    def resource_profile_summary(self, obj):
        profile = obj.resource_profile
        return format_html(
            '<a href="{}">{}</a><span class="tb-source-note">{}</span>',
            reverse(
                "admin:test_platform_testresourceprofile_change",
                args=[profile.pk],
            ),
            profile.name,
            profile.system_id or "-",
        )

    @admin.display(description="测试分类")
    def category_summary(self, obj):
        return _category_badges(obj.test_categories)

    @admin.display(description="审批状态", ordering="status")
    def approval_status(self, obj):
        badge = _status_badge(obj.status, obj.get_status_display())
        if obj.status != ExecutionPlanArtifact.Status.GENERATING:
            return badge
        return format_html(
            "{}{}",
            badge,
            _generation_progress_display(obj, compact=True),
        )

    @admin.display(description="运行状态", ordering="status")
    def execution_status(self, obj):
        if obj.status == ExecutionPlanArtifact.Status.GENERATING:
            return format_html(
                "{}{}",
                _status_badge("pending", "生成中"),
                _generation_progress_display(obj, compact=True),
            )
        if obj.status in {
            ExecutionPlanArtifact.Status.REVIEW,
            ExecutionPlanArtifact.Status.APPROVED,
        }:
            return _status_badge("pending", "等待运行")
        return _status_badge("failed", "需重新生成")

    @admin.display(description="生成进度")
    def generation_progress_display(self, obj):
        return _generation_progress_display(obj)

    def get_record_commands(self, request, obj):
        if obj is None or obj.status == ExecutionPlanArtifact.Status.GENERATING:
            return []
        commands = []
        if obj.status == ExecutionPlanArtifact.Status.REVIEW:
            validation = (obj.compilation_result or {}).get("validation") or {}
            artifacts = (obj.compilation_result or {}).get("artifacts") or []
            if validation.get("passed") is True and artifacts:
                commands.append(
                    {
                        "name": "_approve_execution_plan",
                        "label": "运行",
                        "handler": "run_execution_plan",
                        "kind": "primary",
                    }
                )
        elif obj.status == ExecutionPlanArtifact.Status.APPROVED:
            commands.append(
                {
                    "name": "_run_execution_plan",
                    "label": "运行",
                    "handler": "run_execution_plan",
                    "kind": "primary",
                }
            )
        if obj.status in {
            ExecutionPlanArtifact.Status.REVIEW,
            ExecutionPlanArtifact.Status.BLOCKED,
            ExecutionPlanArtifact.Status.ERROR,
            ExecutionPlanArtifact.Status.APPROVED,
        }:
            commands.append({
                "name": "_revise_execution_plan",
                "label": "按修改需求重新生成当前计划",
                "handler": "run_revise_execution_plan",
                "kind": "secondary",
            })
        return commands

    def show_record_save(self, request, obj) -> bool:
        return bool(
            obj
            and obj.status
            in {
                ExecutionPlanArtifact.Status.REVIEW,
                ExecutionPlanArtifact.Status.BLOCKED,
                ExecutionPlanArtifact.Status.ERROR,
                ExecutionPlanArtifact.Status.APPROVED,
            }
        )

    def get_readonly_fields(self, request, obj=None):
        fields = list(super().get_readonly_fields(request, obj))
        if obj and not self.show_record_save(request, obj):
            fields.append("review_comments")
        return tuple(dict.fromkeys(fields))

    def run_execution_plan(self, request, obj):
        if obj.status == ExecutionPlanArtifact.Status.REVIEW:
            self.approve_execution_plan(request, self.model.objects.filter(pk=obj.pk))
        obj.refresh_from_db(fields=("status",))
        if obj.status == ExecutionPlanArtifact.Status.APPROVED:
            self._queue_run(request, obj)
            return HttpResponseRedirect(
                reverse("admin:test_platform_testexecutionrun_changelist")
            )
        return HttpResponseRedirect(
            reverse("admin:test_platform_executionplanartifact_change", args=[obj.pk])
        )

    def run_revise_execution_plan(self, request, obj):
        comments = str(obj.review_comments or "").strip()
        if not comments:
            self.message_user(request, "请先填写修改要求，再重新生成执行计划。", messages.ERROR)
            return HttpResponseRedirect(
                reverse("admin:test_platform_executionplanartifact_change", args=[obj.pk])
            )

        try:
            queue_execution_plan_generation(
                obj.source_test_plan,
                feedback=comments,
                execution_input=obj.execution_input,
            )
        except Exception as exc:
            self.message_user(request, f"重新生成失败：{exc}", messages.ERROR)
        else:
            if obj.status != ExecutionPlanArtifact.Status.APPROVED:
                obj.status = ExecutionPlanArtifact.Status.CHANGES
                obj.save(update_fields=("status", "updated_at"))
            self.message_user(request, "已按修改需求重新生成执行计划。", messages.INFO)
        return HttpResponseRedirect(
            reverse("admin:test_platform_executionplanartifact_changelist")
        )

    def _queue_run(self, request, artifact):
        from .execution_service import queue_execution_plan_artifact

        try:
            run = queue_execution_plan_artifact(artifact)
        except Exception as exc:
            self.message_user(request, f"“{artifact.title}”运行提交失败：{exc}", messages.ERROR)
            return None
        self.message_user(
            request,
            f"“{artifact.title}”已提交运行，批次 {run.run_id}。",
            messages.SUCCESS,
        )
        return run

    @admin.action(description="批量运行所选执行计划")
    def run_selected(self, request, queryset):
        for artifact in queryset.select_related("source_test_plan", "resource_profile"):
            if artifact.status == ExecutionPlanArtifact.Status.REVIEW:
                self.approve_execution_plan(
                    request,
                    self.model.objects.filter(pk=artifact.pk),
                )
                artifact.refresh_from_db(fields=("status",))
            if artifact.status != ExecutionPlanArtifact.Status.APPROVED:
                self.message_user(
                    request,
                    f"“{artifact.title}”当前不可运行。",
                    messages.WARNING,
                )
                continue
            self._queue_run(request, artifact)

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("source_test_plan__source_intent")
        )

    @admin.display(description="执行计划概览")
    def execution_plan_preview(self, obj):
        plan = (obj.compilation_result or {}).get("plan") or (
            (obj.approved_bundle or {}).get("plan") or {}
        )
        return TestWorkflowAdmin.plan_review_preview(
            self,
            _artifact_preview_proxy(plan=plan),
        )

    @admin.display(description="已冻结运行输入")
    def execution_input_summary(self, obj):
        payload = dict(obj.execution_input or {})
        variables = payload.get("variables") or {}
        if not variables:
            return ""
        variable_rows = format_html_join(
            "",
            "<tr><th>{}</th><td><code>{}</code></td></tr>",
            (
                (
                    name,
                    value
                    if isinstance(value, str)
                    else json.dumps(value, ensure_ascii=False, sort_keys=True),
                )
                for name, value in variables.items()
            ),
        )
        return format_html(
            "<div class='tb-runtime-input-summary'><table><tbody>{}</tbody></table></div>",
            variable_rows,
        )

    @admin.display(description="审批重点")
    def approval_focus_preview(self, obj):
        plan = (obj.compilation_result or {}).get("plan") or (
            (obj.approved_bundle or {}).get("plan") or {}
        )
        sections = []
        for flow in plan.get("flows") or []:
            stages = sorted(
                flow.get("stages") or [],
                key=lambda value: (value.get("order") or 0, value.get("stage_id") or ""),
            )
            stage_sections = []
            for stage in stages:
                if not _stage_has_operations(stage):
                    continue
                executor = str(
                    stage.get("executor_kind")
                    or (stage.get("execution") or {}).get("kind")
                    or ""
                )
                category = EXECUTOR_CATEGORY.get(executor, executor)
                stage_sections.append(
                    format_html(
                        "<section class='tb-approval-stage'>"
                        "<h4><span>{}. {}</span><code>{}</code></h4>"
                        "{}{}"
                        "</section>",
                        stage.get("order") or "?",
                        TEST_CATEGORY_LABELS.get(category, category or "未知测试"),
                        stage.get("stage_id") or "-",
                        self._stage_test_content(stage),
                        self._stage_key_code(stage),
                    )
                )
            if not stage_sections:
                continue
            sections.append(
                format_html(
                    "<section class='tb-approval-flow'><h3>{}</h3>{}</section>",
                    flow.get("name") or flow.get("flow_id") or "未命名流程",
                    format_html_join("", "{}", ((item,) for item in stage_sections)),
                )
            )
        if not sections:
            return "尚无可审批的具体执行内容"
        return format_html(
            "<div class='tb-approval-focus'>"
            "<p>请重点确认执行顺序、实际请求或 SQL、负载方式以及通过标准。</p>"
            "{}</div>",
            format_html_join("", "{}", ((item,) for item in sections)),
        )

    @admin.display(description="网页 Agent 资产来源")
    def page_execution_basis(self, obj):
        """Show the Agent UI content frozen into this approved plan."""

        agent_profiles = (obj.catalog_snapshot or {}).get("agent_ui_profiles") or []
        if agent_profiles:
            plan = (getattr(obj, "compilation_result", None) or {}).get("plan") or {}
            used_refs = {
                str(row.get("operation_ref") or "")
                for flow in plan.get("flows") or []
                for stage in flow.get("stages") or []
                if (stage.get("execution") or {}).get("kind") == "stagehand_agent"
                for row in (stage.get("execution") or {}).get("rows") or []
            }
            cards = []
            for profile in agent_profiles:
                features = [
                    item.get("description") or item.get("operation_ref")
                    for item in profile.get("operations") or []
                    if item.get("operation_ref") in used_refs
                ]
                if not features:
                    continue
                cards.append(
                    format_html(
                        "<section class='tb-capability-card'><h4>网页 Agent 资产</h4>"
                        "<p>起始 URL：<code>{}</code><br>最大步数：{}</p>"
                        "<div><strong>本次功能范围</strong><ul>{}</ul></div></section>",
                        profile.get("start_url") or "-",
                        profile.get("max_steps") or "-",
                        format_html_join("", "<li>{}</li>", ((item,) for item in features)),
                    )
                )
            if cards:
                return format_html(
                    "<details class='tb-capability-source'><summary>查看网页 Agent 资产</summary>"
                    "<div class='tb-knowledge-list'>{}</div></details>",
                    format_html_join("", "{}", ((card,) for card in cards)),
                )
        return "本计划不包含页面操作"

    @admin.display(description="规则校验")
    def validation_summary(self, obj):
        validation = (obj.compilation_result or {}).get("validation") or {}
        if obj.status == ExecutionPlanArtifact.Status.GENERATING:
            return "后台正在生成并校验执行计划"
        findings = validation.get("findings") or []
        if validation.get("passed") is True:
            return _status_badge("approved", "全部规则通过")
        if not findings:
            return "尚无校验结果"
        return format_html(
            "<div class='tb-validation-issues'><strong>需要处理的问题</strong>{}</div>",
            format_html_join(
                "",
                "<p>{}</p>",
                (
                    (item.get("message") or "计划内容需要调整",)
                    for item in findings
                ),
            ),
        )

    @admin.display(description="各分类生成的执行文件")
    def execution_artifact_preview(self, obj):
        artifacts = (obj.compilation_result or {}).get("artifacts") or []
        if not artifacts:
            return "尚未生成执行文件"
        cards = []
        artifact_counts = {}
        for artifact in artifacts:
            executor = str(artifact.get("executor_kind") or "")
            category = EXECUTOR_CATEGORY.get(executor, executor)
            artifact_counts[category] = artifact_counts.get(category, 0) + 1
            refs = artifact.get("artifact_refs") or []
            files = [
                (
                    f"{item.get('kind')}: {item.get('path_ref')} "
                    f"({str(item.get('sha256') or '')[:15]}…)"
                )
                for item in refs
                if item.get("kind") != "manifest"
            ]
            details = [
                self._execution_artifact_detail(obj, artifact, ref)
                for ref in refs
                if ref.get("kind") != "manifest"
            ]
            cards.append(
                format_html(
                    "<details class='tb-artifact-stage'><summary>"
                    "<span class='tb-artifact-stage-title'><strong>{}</strong>"
                    "<code>{}</code></span>"
                    "<span class='tb-artifact-stage-files'>{}</span>"
                    "<span class='tb-artifact-stage-action'>查看生成文件</span>"
                    "</summary><div class='tb-artifact-stage-body'>"
                    "<div class='tb-generated-files'>{}</div></div></details>",
                    TEST_CATEGORY_LABELS.get(category, category),
                    artifact.get("stage_id") or "-",
                    _display_lines(files),
                    format_html_join("", "{}", ((item,) for item in details)),
                )
            )
        return format_html(
            "<div class='tb-artifact-summary'><p class='tb-scope-line'>"
            "<strong>执行内容：</strong>{}</p>"
            "<div class='tb-artifact-stage-list'>{}</div></div>",
            "、".join(
                f"{TEST_CATEGORY_LABELS.get(category, category)} {count} 组"
                for category, count in artifact_counts.items()
            ),
            format_html_join("", "{}", ((card,) for card in cards)),
        )

    @staticmethod
    def _json_text(value):
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        )

    @classmethod
    def _assertion_text(cls, assertion):
        assertion_kind = assertion.get("kind")
        location = (
            assertion.get("column")
            or assertion.get("path")
            or assertion.get("name")
            or assertion.get("metric")
            or assertion_kind
            or "结果"
        )
        operator = assertion.get("operator") or "检查"
        expected = assertion.get("expected")
        expected_text = (
            ""
            if expected is None
            else " " + cls._json_text(expected)
        )
        unit = f" {assertion.get('unit')}" if assertion.get("unit") else ""
        statement = str(assertion.get("statement") or "").strip()
        technical = f"{location} {operator}{expected_text}{unit}"
        return f"{statement}（{technical}）" if statement else technical

    @classmethod
    def _stage_key_code(cls, stage):
        """Render the key executable operations and assertions for review."""

        execution = (stage or {}).get("execution") or {}
        kind = execution.get("kind")
        lines = []

        def quoted(value):
            return cls._json_text(value)

        def assertion_line(assertion, actual):
            operator = str(assertion.get("operator") or "check")
            symbols = {
                "equals": "==",
                "not_equals": "!=",
                "gt": ">",
                "gte": ">=",
                "lt": "<",
                "lte": "<=",
                "contains": "contains",
                "exists": "exists",
                "not_exists": "not exists",
                "not_null": "is not null",
                "null": "is null",
            }
            expected = assertion.get("expected")
            expression = f"{actual} {symbols.get(operator, operator)}"
            if expected is not None:
                expression += f" {quoted(expected)}"
            return "assert " + expression

        if kind == "stagehand_agent":
            lines.append(f"open({quoted(execution.get('start_url'))})")
            lines.append(f"max_steps = {execution.get('max_steps')}")
            for index, item in enumerate(execution.get("rows") or [], start=1):
                lines.append(f"action_{index} = agent_execute({quoted(item.get('action'))})")
                for assertion in item.get("assertions") or []:
                    statement = assertion.get("statement") or "页面检查"
                    lines.append(f"check_{index} = verify({quoted(statement)})")
        elif kind == "http_api":
            for index, item in enumerate(execution.get("requests") or [], start=1):
                lines.append(
                    f"response_{index} = http_request("
                    f"{quoted(item.get('method'))}, {quoted(item.get('path'))})"
                )
                for assertion in item.get("assertions") or []:
                    actual = {
                        "status": f"response_{index}.status_code",
                        "body_contains": f"response_{index}.text",
                        "json": f"response_{index}.json",
                        "header": f"response_{index}.headers",
                    }.get(
                        assertion.get("kind"),
                        f"response_{index}.{assertion.get('kind') or 'result'}",
                    )
                    lines.append(assertion_line(assertion, actual))
        elif kind == "database":
            for index, item in enumerate(execution.get("operations") or [], start=1):
                if item.get("sql"):
                    parameters = item.get("parameters_refs") or {}
                    if parameters:
                        lines.append(
                            f"params_{index} = {quoted(parameters)}"
                        )
                    lines.append(f"result_{index} = db.query(\"\"\"")
                    sql_lines = str(item["sql"]).strip().splitlines()
                    lines.extend(sql_lines)
                    lines.append(
                        f"\"\"\", params=params_{index})"
                        if parameters
                        else "\"\"\")"
                    )
                else:
                    lines.append(
                        f"result_{index} = run_imported_query("
                        f"{quoted(item.get('operation_ref'))})"
                    )
                for assertion in item.get("assertions") or []:
                    if assertion.get("kind") == "row_count":
                        actual = f"result_{index}.row_count"
                    elif assertion.get("kind") == "exists":
                        actual = f"bool(result_{index}.rows)"
                    else:
                        actual = (
                            f"result_{index}[{quoted(assertion.get('column'))}]"
                        )
                    lines.append(assertion_line(assertion, actual))
        elif kind == "performance":
            load_stages = execution.get("stages") or []
            for index, load in enumerate(load_stages, start=1):
                lines.append(
                    f"load_{index} = run_load("
                    f"users={load.get('virtual_users')}, "
                    f"duration_seconds={load.get('duration_seconds')})"
                )
            for threshold in execution.get("thresholds") or []:
                lines.append(
                    assertion_line(
                        {
                            **threshold,
                            "expected": threshold.get("value"),
                        },
                        f"metrics[{quoted(threshold.get('metric'))}]",
                    )
                )
        elif kind == "tcp_port":
            for index, item in enumerate(execution.get("probes") or [], start=1):
                lines.append(
                    f"probe_{index} = tcp_connect("
                    f"host={item.get('host_ref')}, port={item.get('port')}, "
                    f"timeout={item.get('timeout_seconds')})"
                )
                for assertion in item.get("assertions") or []:
                    actual = (
                        f"probe_{index}.latency_ms"
                        if assertion.get("kind") == "connect_latency_ms"
                        else f"probe_{index}.state"
                    )
                    lines.append(assertion_line(assertion, actual))

        if not lines:
            return format_html(
                "<div class='tb-key-code'><h4>关键执行逻辑</h4>"
                "<p class='tb-capability-warning'>没有可展示的执行逻辑。</p></div>"
            )
        return format_html(
            "<section class='tb-key-code'><h4>关键执行逻辑</h4>"
            "<pre>{}</pre></section>",
            "\n".join(lines),
        )

    @classmethod
    def _stage_test_content(cls, stage):
        execution = (stage or {}).get("execution") or {}
        kind = execution.get("kind")
        rows = []
        headers = []
        if kind == "stagehand_agent":
            headers = ["步骤", "Agent 页面操作", "检查"]
            for index, item in enumerate(execution.get("rows") or [], start=1):
                rows.append(
                    [
                        index,
                        item.get("action") or item.get("operation_ref") or "-",
                        "；".join(
                            cls._assertion_text(value)
                            for value in item.get("assertions") or []
                        ) or "无独立检查",
                    ]
                )
        elif kind == "http_api":
            headers = ["请求", "用途", "输入绑定", "检查"]
            for item in execution.get("requests") or []:
                bindings = [
                    cls._json_text(value.get("input_refs") or {})
                    for value in item.get("data_bindings") or []
                ]
                rows.append(
                    [
                        f"{item.get('method')} {item.get('path')}",
                        item.get("action") or "-",
                        "；".join(bindings) or "无",
                        "；".join(
                            cls._assertion_text(value)
                            for value in item.get("assertions") or []
                        )
                        or "无",
                    ]
                )
        elif kind == "database":
            headers = ["SQL 来源", "实际 SQL/查询", "运行参数", "检查"]
            origin_labels = {
                "catalog": "导入计划中的查询引用",
                "knowledge_reused": "复用业务知识库 SQL",
                "ai_generated": "AI 新生成，等待人工审批",
            }
            for item in execution.get("operations") or []:
                sql_origin = item.get("sql_origin") or "catalog"
                operation_ref = str(item.get("operation_ref") or "")
                rows.append(
                    [
                        origin_labels.get(sql_origin, sql_origin),
                        item.get("sql") or operation_ref or "-",
                        cls._json_text(item.get("parameters_refs") or {})
                        if item.get("parameters_refs")
                        else "无",
                        "；".join(
                            cls._assertion_text(value)
                            for value in item.get("assertions") or []
                        )
                        or "无",
                    ]
                )
        elif kind == "performance":
            headers = ["负载阶段", "执行目标", "输入绑定", "通过标准"]
            load_stages = "；".join(
                f"{value.get('virtual_users')} 用户 / {value.get('duration_seconds')} 秒"
                for value in execution.get("stages") or []
            )
            thresholds = "；".join(
                cls._assertion_text(
                    {
                        **value,
                        "expected": value.get("value"),
                        "kind": "threshold",
                    }
                )
                for value in execution.get("thresholds") or []
            )
            rows.append(
                [
                    load_stages or "未配置",
                    execution.get("profile_ref")
                    or execution.get("driver_ref")
                    or "-",
                    "；".join(
                        cls._json_text(value.get("input_refs") or {})
                        for value in execution.get("data_bindings") or []
                    )
                    or "无",
                    thresholds or "无",
                ]
            )
        elif kind == "tcp_port":
            headers = ["探测目标", "超时", "用途", "检查"]
            for item in execution.get("probes") or []:
                rows.append(
                    [
                        f"{item.get('host_ref')}:{item.get('port')}",
                        f"{item.get('timeout_seconds')} 秒",
                        item.get("action") or "TCP 连通性探测",
                        "；".join(
                            cls._assertion_text(value)
                            for value in item.get("assertions") or []
                        )
                        or "无",
                    ]
                )
        if not headers or not rows:
            return format_html(
                "<p class='tb-capability-warning'>尚无可展示的具体测试内容。</p>"
            )
        return format_html(
            "<div class='tb-semantic-preview'><table><thead><tr>{}</tr></thead>"
            "<tbody>{}</tbody></table></div>",
            format_html_join("", "<th>{}</th>", ((item,) for item in headers)),
            format_html_join(
                "",
                "<tr>{}</tr>",
                (
                    (
                        format_html_join(
                            "",
                            "<td>{}</td>",
                            ((value,) for value in row),
                        ),
                    )
                    for row in rows
                ),
            ),
        )

    def _execution_artifact_path(self, obj, artifact, ref):
        from .planning.artifact_paths import generated_files_root

        root = Path(settings.TEST_PLATFORM_ARTIFACT_ROOT).resolve()
        artifact_root = (root / Path(obj.artifact_root_ref)).resolve()
        artifact_root.relative_to(root)
        stage_root = (
            generated_files_root(artifact_root, artifact.get("executor_kind"))
            / obj.plan_id
            / f"v{obj.version}"
            / str(artifact.get("flow_id"))
            / str(artifact.get("stage_id"))
        ).resolve()
        if not stage_root.is_dir():
            stage_root = (
                artifact_root
                / obj.plan_id
                / f"v{obj.version}"
                / str(artifact.get("flow_id"))
                / str(artifact.get("stage_id"))
            ).resolve()
        stage_root.relative_to(artifact_root)
        source_path = (stage_root / str(ref.get("path_ref") or "")).resolve()
        source_path.relative_to(stage_root)
        return source_path

    def _execution_artifact_detail(self, obj, artifact, ref):
        kind = str(ref.get("kind") or "")
        label = {
            "workbook": "查看 UI 测试步骤",
            "payload": "查看执行配置",
            "pytest_source": "查看完整 pytest 代码",
        }.get(kind, "查看文件内容")
        try:
            source_path = self._execution_artifact_path(obj, artifact, ref)
            if kind == "workbook":
                content = self._workbook_preview(source_path)
            else:
                raw = source_path.read_text(encoding="utf-8")
                if len(raw) > 1_000_000:
                    raw = raw[:1_000_000] + "\n\n内容过长，页面仅显示前 1 MB。"
                if kind == "payload" or source_path.suffix.lower() == ".json":
                    try:
                        raw = json.dumps(
                            json.loads(raw),
                            ensure_ascii=False,
                            indent=2,
                        )
                    except json.JSONDecodeError:
                        pass
                content = format_html("<pre class='tb-source-preview'>{}</pre>", raw)
        except (OSError, UnicodeError, ValueError) as exc:
            content = format_html(
                "<p class='tb-capability-warning'>无法读取生成文件：{}</p>",
                str(exc) or "文件不存在",
            )
        download_url = reverse(
            f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_execution_artifact",
            args=[
                obj.pk,
                artifact.get("flow_id"),
                artifact.get("stage_id"),
                ref.get("path_ref"),
            ],
        )
        return format_html(
            "<section class='tb-execution-detail'>"
            "<header><h4>{}</h4><code>{}</code>"
            "<a class='tb-audit-download' href='{}'>下载原始文件</a></header>"
            "{}</section>",
            label,
            ref.get("path_ref") or "-",
            download_url,
            content,
        )

    @staticmethod
    def _workbook_preview(source_path):
        from openpyxl import load_workbook

        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Case"] if "Case" in workbook.sheetnames else workbook.active
            max_row = sheet.max_row
            values = list(sheet.iter_rows(values_only=True, max_row=201))
        finally:
            workbook.close()
        if not values:
            return format_html("<p>执行表为空</p>")
        headers = [str(value or "-") for value in values[0]]
        header_html = format_html_join("", "<th>{}</th>", ((item,) for item in headers))
        row_html = []
        for row in values[1:]:
            cells = list(row) + [""] * max(0, len(headers) - len(row))
            row_html.append(
                format_html(
                    "<tr>{}</tr>",
                    format_html_join(
                        "",
                        "<td>{}</td>",
                        ((str(value or ""),) for value in cells[:len(headers)]),
                    ),
                )
            )
        suffix = (
            format_html("<p>页面仅显示前 200 条步骤。</p>")
            if max_row and max_row > 201
            else ""
        )
        return format_html(
            "<div class='tb-workbook-preview'><table><thead><tr>{}</tr></thead>"
            "<tbody>{}</tbody></table>{}</div>",
            header_html,
            format_html_join("", "{}", ((row,) for row in row_html)),
            suffix,
        )

    def download_execution_artifact(
        self,
        request,
        object_id,
        flow_id,
        stage_id,
        path_ref,
    ):
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("执行计划不存在")
        if not self.has_view_or_change_permission(request, obj):
            raise PermissionDenied
        artifact = next(
            (
                item
                for item in (obj.compilation_result or {}).get("artifacts") or []
                if str(item.get("flow_id")) == flow_id
                and str(item.get("stage_id")) == stage_id
            ),
            None,
        )
        if artifact is None:
            raise Http404("执行阶段不存在")
        ref = next(
            (
                item
                for item in artifact.get("artifact_refs") or []
                if item.get("kind") != "manifest"
                and str(item.get("path_ref")) == path_ref
            ),
            None,
        )
        if ref is None:
            raise Http404("执行文件不存在")
        try:
            source_path = self._execution_artifact_path(obj, artifact, ref)
        except (OSError, ValueError) as exc:
            raise Http404("执行文件路径无效") from exc
        if not source_path.is_file():
            raise Http404("执行文件不存在")
        content_type = mimetypes.guess_type(source_path.name)[0] or "application/octet-stream"
        return FileResponse(
            source_path.open("rb"),
            as_attachment=True,
            filename=source_path.name,
            content_type=content_type,
        )

    def _review(self, request, queryset, decision):
        from .approval_service import persist_execution_plan_review
        from .planning.catalogs import PlanningCatalogSnapshot
        from .service_factory import get_workflow

        for artifact in queryset:
            try:
                if artifact.status not in {
                    ExecutionPlanArtifact.Status.REVIEW,
                    ExecutionPlanArtifact.Status.BLOCKED,
                }:
                    raise ValidationError("只有待审批或已阻塞执行计划可以处理")
                output = get_workflow().review_plan(
                    _compilation_result_from_payload(artifact.compilation_result),
                    PlanningCatalogSnapshot.model_validate(artifact.catalog_snapshot),
                    decision=decision,
                    comments=artifact.review_comments or (
                        "后台审核通过" if decision.value == "approved" else "请补充审批意见"
                    ),
                )
                artifact = persist_execution_plan_review(
                    artifact.pk,
                    expected_status=artifact.status,
                    review_payload=_json(output.review),
                    approved_bundle=(
                        _json(output.approved_bundle)
                        if output.approved_bundle is not None
                        else None
                    ),
                )
                self.message_user(
                    request,
                    f"“{artifact.title}”{artifact.get_status_display()}。",
                    messages.SUCCESS,
                )
            except Exception as exc:
                if not isinstance(exc, ValidationError):
                    ExecutionPlanArtifact.objects.filter(pk=artifact.pk).update(
                        last_error=str(exc)[:20_000]
                    )
                self.message_user(request, f"“{artifact.title}”审批失败：{exc}", messages.ERROR)

    @admin.action(description="审批通过执行计划")
    def approve_execution_plan(self, request, queryset):
        from .planning.contracts import PlanReviewDecision

        self._review(request, queryset, PlanReviewDecision.APPROVED)

_WORKFLOW_STEP_ORDER = {
    "TestResourceProfile": 10,
    "ApprovedKnowledgeEntry": 15,
    "TestIntentImport": 20,
    "TestPlanArtifact": 30,
    "ExecutionPlanArtifact": 40,
    "TestExecutionRun": 50,
}

_default_get_app_list = admin.site.get_app_list


def _ordered_workflow_app_list(self, request, app_label=None):
    app_list = _default_get_app_list(request, app_label)
    visible = []
    for app in app_list:
        if app["app_label"] == "auth":
            # Django Admin imports auth models internally; the local workspace
            # does not expose them as business modules.
            continue
        if app["app_label"] != "test_platform":
            continue
        app["name"] = "测试流程"
        app["models"] = sorted(
            app["models"],
            key=lambda item: _WORKFLOW_STEP_ORDER.get(item["object_name"], 999),
        )
        visible.append(app)
    return visible


admin.site.get_app_list = _ordered_workflow_app_list.__get__(admin.site, type(admin.site))


@admin.register(TestExecutionRun)
class TestExecutionRunAdmin(SingleRecordActionAdmin, admin.ModelAdmin):
    actions = ("mark_selected", "unmark_selected", "delete_selected")
    change_list_template = "admin/test_platform/execution_run_change_list.html"
    list_display = (
        "marked_status",
        "run_overview",
        "pass_result",
        "started_at_display",
        "run_source_summary",
        "quick_report_link",
    )
    list_filter = (
        MarkedRecordFilter,
        ExecutionRunDateFilter,
        ExecutionRunStatusFilter,
    )
    search_fields = ("run_id", "plan_id", "design_id", "target_system_id")
    list_display_links = ("run_overview",)
    readonly_fields = (
        "unified_status",
        "category_display",
        "category_result_summary",
        "error_summary",
        "report_links",
    )
    fieldsets = (
        (
            "执行结果",
            {
                "fields": (
                    "run_id",
                    "unified_status",
                    "category_display",
                    "category_result_summary",
                    "error_summary",
                    "report_links",
                )
            },
        ),
        (
            "批次信息",
            {
                "fields": (
                    "execution_plan",
                    "resource_profile",
                    "started_at",
                    "finished_at",
                    "duration_ms",
                )
            },
        ),
    )

    failure_statuses = frozenset(
        {
            TestExecutionRun.Status.FAILED,
            TestExecutionRun.Status.BLOCKED,
            TestExecutionRun.Status.ERROR,
            TestExecutionRun.Status.INCONCLUSIVE,
        }
    )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            "execution_plan__source_test_plan__source_intent",
            "resource_profile",
        )

    @admin.display(description="运行批次号", ordering="run_id")
    def run_id_display(self, obj):
        return format_html(
            '<span class="tb-run-id" title="{}">{}</span>',
            obj.run_id,
            obj.run_id,
        )

    @admin.display(description="测试内容")
    def run_overview(self, obj):
        plan = obj.execution_plan.source_test_plan if obj.execution_plan_id else None
        intent = plan.source_intent if plan is not None and plan.source_intent_id else None
        title = (
            getattr(intent, "title", "")
            or getattr(plan, "title", "")
            or getattr(obj.execution_plan, "title", "")
            or "未命名测试"
        )
        short_id = str(obj.run_id or "")[-8:] or "-"
        return format_html(
            "<div class='tb-run-overview'><strong>{}</strong><span>{}</span>"
            "<small>记录 {}</small></div>",
            title,
            self.category_display(obj),
            short_id,
        )

    @admin.display(description="开始时间", ordering="started_at")
    def started_at_display(self, obj):
        if obj.started_at is None:
            return "-"
        return timezone.localtime(obj.started_at).strftime("%Y-%m-%d %H:%M")

    @admin.display(description="测试分类")
    def category_display(self, obj):
        categories = []
        if obj.execution_plan_id:
            categories = obj.execution_plan.test_categories
        if not categories:
            categories = list(((obj.result_summary or {}).get("categories") or {}).keys())
        return _category_badges(categories)

    @admin.display(description="各分类执行结果")
    def category_result_summary(self, obj):
        categories = (obj.result_summary or {}).get("categories") or {}
        if not categories:
            return "旧记录未保存分类汇总，请查看报告中的执行阶段"
        rows = []
        for category, counts in categories.items():
            failed_count = sum(
                counts.get(status, 0)
                for status in ("failed", "error", "inconclusive")
            )
            display_counts = (
                ("成功", counts.get("passed", 0)),
                ("失败", failed_count),
                ("阻断", counts.get("blocked", 0)),
                ("预检", counts.get("dry_run", 0)),
            )
            statuses = "、".join(
                f"{label} {count}"
                for label, count in display_counts
                if count
            ) or "无结果"
            rows.append(
                format_html(
                    "<tr><th>{}</th><td>{}</td><td>{}</td></tr>",
                    TEST_CATEGORY_LABELS.get(category, category),
                    counts.get("total", 0),
                    statuses,
                )
            )
        return format_html(
            "<div class='tb-category-results'><table><thead><tr>"
            "<th>测试分类</th><th>执行阶段</th><th>结果</th></tr></thead>"
            "<tbody>{}</tbody></table></div>",
            format_html_join("", "{}", ((row,) for row in rows)),
        )

    @admin.display(description="执行状态", ordering="status")
    def pass_result(self, obj):
        if obj.status == TestExecutionRun.Status.PASSED:
            return _status_badge("passed", "成功")
        if obj.status in self.failure_statuses:
            if obj.status == TestExecutionRun.Status.BLOCKED:
                return _status_badge("blocked", "阻断")
            return _status_badge("failed", "失败")
        if obj.status == TestExecutionRun.Status.DRY_RUN:
            return _status_badge("pending", "预检")
        if obj.status == TestExecutionRun.Status.QUEUED:
            return _status_badge("pending", "等待中")
        return _status_badge("running", "执行中")

    @admin.display(description="执行状态", ordering="status")
    def unified_status(self, obj):
        return self.pass_result(obj)

    @admin.display(description="报告")
    def report_links(self, obj):
        if not obj.report_paths:
            return "-"
        return format_html(
            " ".join(
                '<a href="{}">{}</a>'.format(
                    reverse("test_platform_report", kwargs={"run_id": obj.run_id, "kind": key}),
                    key,
                )
                for key in ("html", "json", "junit")
                if key in obj.report_paths
            )
        )

    @admin.display(description="快速报告")
    def quick_report_link(self, obj):
        paths = obj.report_paths or {}
        kind = next((value for value in ("html", "json", "junit") if value in paths), None)
        if kind is None:
            if obj.report_status == TestExecutionRun.ReportStatus.PENDING:
                return "生成中"
            return "-"
        labels = {"html": "查看报告", "json": "查看 JSON", "junit": "查看 JUnit"}
        return format_html(
            '<a class="button" href="{}">{}</a>',
            reverse("test_platform_report", kwargs={"run_id": obj.run_id, "kind": kind}),
            labels[kind],
        )

    @admin.display(description="测试计划")
    def test_plan_link(self, obj):
        plan = obj.execution_plan.source_test_plan if obj.execution_plan_id else None
        if plan is None:
            return obj.design_id or "-"
        return format_html(
            '<a href="{}">{}</a>',
            reverse("admin:test_platform_testplanartifact_change", args=[plan.pk]),
            plan.title or plan.artifact_id,
        )

    @admin.display(description="执行计划")
    def execution_plan_link(self, obj):
        if not obj.execution_plan_id:
            return obj.plan_id or "-"
        return format_html(
            '<a href="{}">{}</a>',
            reverse(
                "admin:test_platform_executionplanartifact_change",
                args=[obj.execution_plan_id],
            ),
            obj.execution_plan.title or obj.execution_plan.artifact_id,
        )

    @admin.display(description="测试需求")
    def requirement_link(self, obj):
        plan = obj.execution_plan.source_test_plan if obj.execution_plan_id else None
        intent = plan.source_intent if plan is not None and plan.source_intent_id else None
        if plan is not None and plan.source_kind == TestPlanArtifact.SourceKind.IMPORTED:
            return format_html(
                '<span class="tb-source-kind tb-source-kind--imported">'
                "手动导入测试计划</span>"
            )
        if intent is None:
            return "-"
        return format_html(
            '<a href="{}">{}</a>',
            reverse("admin:test_platform_testintentimport_change", args=[intent.pk]),
            intent.title or intent.workflow_id,
        )

    @admin.display(description="来源")
    def run_source_summary(self, obj):
        return format_html(
            "<div class='tb-run-source'><span><b>测试意图</b>{}</span>"
            "<span><b>测试计划</b>{}</span><span><b>执行计划</b>{}</span></div>",
            self.requirement_link(obj),
            self.test_plan_link(obj),
            self.execution_plan_link(obj),
        )

    @admin.display(description="错误与限制")
    def error_summary(self, obj):
        errors = [str(value) for value in (obj.errors or []) if str(value).strip()]
        limitations = [
            str(value)
            for value in ((obj.result_summary or {}).get("limitations") or [])
            if str(value).strip()
        ]
        if not errors and not limitations:
            return "无"
        rows = [f"错误：{value}" for value in errors]
        rows.extend(f"限制：{value}" for value in limitations)
        return _display_lines(rows)

    def get_readonly_fields(self, request, obj=None):
        return tuple(field.name for field in self.model._meta.fields) + self.readonly_fields

    def has_add_permission(self, request):
        return False

    def show_record_save(self, request, obj) -> bool:
        return False

    def has_change_permission(self, request, obj=None):
        return True

    def has_delete_permission(self, request, obj=None):
        return True

    def changelist_view(self, request, extra_context=None):
        params = request.GET.copy()
        params.pop("date_from", None)
        params.pop("date_to", None)
        params.pop("p", None)
        context = {
            "title": "执行历史",
            "page_description": "每次运行的测试内容、结果、来源、报告与失败定位",
            "run_date_form": ExecutionRunDateFilterForm(
                {
                    "date_from": request.GET.get("date_from", ""),
                    "date_to": request.GET.get("date_to", ""),
                }
            ),
            "run_date_clear_query": params.urlencode(),
        }
        context.update(extra_context or {})
        return super().changelist_view(request, context)


__all__ = [
    "ExecutionPlanArtifactAdmin",
    "TestResourceProfileAdmin",
    "TestExecutionRunAdmin",
    "TestIntentImportAdmin",
    "TestPlanArtifactAdmin",
    "TestWorkflowAdmin",
    "TestWorkflowForm",
]
